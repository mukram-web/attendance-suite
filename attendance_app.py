"""
Be10X — AI CAP Attendance (unified app)
========================================

The app renders a PREBUILT store when one is available: pipeline.py (run by
GitHub Actions every Monday 06:00 IST) fetches the roster + Zoom reports, marks
attendance, and uploads one small attendance.duckdb to a private Drive folder.
The app just downloads that file and renders — opens in seconds, and the
memory-heavy marking never runs on the Streamlit server.

Fallbacks, in order, when no store is configured:
  • legacy live mode — fetch + mark right here (slow first load), or
  • manual upload — roster .xlsx (+ optional L2 + attendee .zip).

Run locally:   streamlit run attendance_app.py
"""
from __future__ import annotations
import hashlib
import hmac
import io
import json
import pickle
import datetime as _dt
from datetime import datetime

import polls as _polls_mod          # the NPS rule, shared with the pipeline
from pathlib import Path

import streamlit as st

import attendance_core as ac
import dashboard_core as dc
import live_data
import data as ddata          # new dashboard data layer (aliased; 'data' is used as a local below)
import sheets as dsheets      # gspread / xlsx source adapter
import dash_view              # Plotly drill-down dashboard UI

st.set_page_config(page_title="Be10X Attendance", page_icon="📊", layout="wide")

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# ───────────────────────── password gate ─────────────────────────────────────
# This app is reachable by anyone with the link, and its Roster tab shows student
# emails and phone numbers. Everything below — including downloading the store —
# happens only after this returns, so an unauthenticated visitor never causes the
# PII to be fetched, let alone rendered.
#
# One SHARED password, read from secrets. It is a deterrent against a stray link,
# not real authentication: it cannot tell users apart, cannot be revoked for one
# person, and is only as private as the least careful person it is given to. For
# genuine access control use Streamlit Cloud's own viewer allow-list
# (Settings -> Sharing), which authenticates against real accounts. See §6.
_PASSWORD_KEY = "app_password"


def _password_ok() -> None:
    """Stop the script unless this session has entered the right password.

    Fails CLOSED: with no password configured nobody gets in. An unset secret is
    far more likely to mean "not set up yet" than "deliberately public", and the
    failure mode of guessing wrong is publishing the roster."""
    if st.session_state.get("_authed"):
        return
    try:
        expected = str(st.secrets.get(_PASSWORD_KEY, "") or "")
    except Exception:            # no secrets file at all
        expected = ""

    st.title("📊 Be10X — AI CAP Attendance")
    if not expected:
        st.error(
            f"**No `{_PASSWORD_KEY}` is configured, so the app is locked.** "
            "Add it to the app's secrets (Streamlit Cloud: Manage app → Settings → "
            f"Secrets) as `{_PASSWORD_KEY} = \"your-password\"`, or to "
            "`.streamlit/secrets.toml` when running locally."
        )
        st.stop()

    def _submit():
        # compare_digest keeps the check constant-time, and the plaintext is
        # dropped from session state the moment it has been used.
        if hmac.compare_digest(st.session_state.get("_pw", ""), expected):
            st.session_state["_authed"] = True
        else:
            st.session_state["_authed"] = False
        st.session_state.pop("_pw", None)

    st.text_input("Password", type="password", key="_pw", on_change=_submit,
                  placeholder="Enter the password and press Enter")
    if st.session_state.get("_authed") is False:
        st.error("Incorrect password.")
    st.caption("Ask the AI CAP team for access.")
    st.stop()


_password_ok()




# ───────────────────────────── cached heavy work ─────────────────────────────
_DISK_CACHE = Path(__file__).parent / ".cache"


def _disk_memo(name: str, key_material: bytes, builder):
    """Persistent twin of st.cache_data: the in-memory cache dies with the server
    process, so a restart used to redo every expensive step (marking 170+ session
    columns takes minutes). Results are pickled to .cache/ keyed by an input hash;
    only one file per artefact is kept (a new key evicts the old one). All disk
    errors fall through to just rebuilding — the cache can never break the app."""
    digest = hashlib.sha256(key_material).hexdigest()[:24]
    path = _DISK_CACHE / f"{name}_{digest}.pkl"
    if path.exists():
        try:
            return pickle.loads(path.read_bytes())
        except Exception:
            pass
    value = builder()
    try:
        _DISK_CACHE.mkdir(exist_ok=True)
        for old in _DISK_CACHE.glob(f"{name}_*.pkl"):    # evict superseded keys
            old.unlink(missing_ok=True)
        tmp = path.with_suffix(".tmp")
        tmp.write_bytes(pickle.dumps(value))
        tmp.replace(path)
    except Exception:
        pass
    return value


def _inputs_key(roster_bytes, l2_bytes, attendee_files, mode) -> bytes:
    h = hashlib.sha256()
    h.update(roster_bytes)
    h.update(l2_bytes or b"")
    for name, data in sorted(attendee_files or ()):
        h.update(name.encode("utf-8", "replace"))
        h.update(data)
    h.update(mode.encode())
    return h.digest()


@st.cache_data(show_spinner="Marking new sessions…")
def _mark(roster_bytes, l2_bytes, attendee_files, mode):
    """Run the marker once per unique input set. values_only=True so formula-based
    attendance columns keep their values through the save (the dashboard reads
    values, not formulas)."""
    return _disk_memo(
        "marked", _inputs_key(roster_bytes, l2_bytes, attendee_files, mode),
        lambda: ac.process_files(roster_bytes, l2_bytes, list(attendee_files),
                                 mode=mode, values_only=True))


@st.cache_data(show_spinner="Building dashboard…")
def _compute(roster_bytes):
    return _disk_memo("compute", hashlib.sha256(roster_bytes).digest(),
                      lambda: dc.compute(roster_bytes))


@st.cache_data(show_spinner=False)
def _attendee_names(nonce):
    """All attendee filenames (one fast Drive query) — only used to label sessions
    with their real topic name via the Webinar-ID join."""
    try:
        return tuple(live_data.list_attendee_names())
    except Exception:
        return ()


@st.cache_data(show_spinner="Building dashboard from the updated roster…")
def _build_dashboard(roster_bytes, attendee_names, l2_bytes):
    """The dashboard reads the SAME updated roster the app produced (the one you
    can download) — no second fetch. Session names come from the Webinar-ID join
    (attendee filenames ⋈ L2)."""
    key = hashlib.sha256()
    key.update(roster_bytes)
    key.update("\n".join(sorted(attendee_names)).encode("utf-8", "replace"))
    key.update(l2_bytes or b"")
    return _disk_memo("dashdata", key.digest(),
                      lambda: _build_dashboard_impl(roster_bytes, attendee_names, l2_bytes))


def _build_dashboard_impl(roster_bytes, attendee_names, l2_bytes):
    import io as _io
    from openpyxl import load_workbook
    wb = load_workbook(_io.BytesIO(roster_bytes), read_only=True, data_only=True)
    tabs = {ws.title: [list(r) for r in ws.iter_rows(values_only=True)]
            for ws in wb.worksheets}
    wb.close()
    topics, l2_labels, mentors = dsheets.webinar_topic_lookup(
        [(n, None) for n in attendee_names], l2_bytes,
        with_labels=True, with_mentors=True)
    DATA, summary = ddata.build(tabs, topics, l2_labels, None, mentors)
    _prepend_intro_sessions(DATA)
    return DATA, summary


def _prepend_intro_sessions(DATA: dict) -> None:
    """First session of every batch = the intro call. Counts come from
    intro_attendance.json (unique attendees per batch, precomputed from the
    'L2 customer' sheet — aggregates only). Batches missing from the file are
    left untouched. Intro rows are flagged is_intro so avg/peak/low keep
    measuring real class sessions."""
    import json
    from pathlib import Path
    f = Path(__file__).parent / "intro_attendance.json"
    if not f.exists():
        return
    intro = json.loads(f.read_text(encoding="utf-8"))
    for code, d in DATA.items():
        att = intro.get(code)
        if not att:
            continue
        stg = d["strength"]
        d["sessions"].insert(0, {
            "col": None, "mm": None,
            "date_lbl": "Intro call",
            "topic": "Intro call (pre-batch)",
            "present": att, "absent": max(0, stg - att), "total": stg,
            "pct": round(min(100.0, 100 * att / stg), 1),
            "present_only": False, "no_l2": False, "is_intro": True,
            # the intro call predates the batch and has no feedback poll
            "rating": None, "rating_n": 0,
        })


# ─────────────────────── prebuilt store (pipeline.py) ────────────────────────
_STORE_PATH = _DISK_CACHE / "attendance.duckdb"


def _store_folder_id() -> str:
    try:
        return st.secrets.get("drive", {}).get("store_folder_id", "")
    except Exception:
        return ""


# The store is re-checked this often, so the Monday 06:00 IST rebuild reaches
# viewers on its own. Without a TTL, one cached load would pin the whole server
# to that week's data until somebody happened to press Refresh.
_STORE_TTL_SECONDS = 30 * 60


@st.cache_data(show_spinner="Loading the latest dashboard…", ttl=_STORE_TTL_SECONDS)
def _load_store(nonce):
    """Download the prebuilt attendance.duckdb the pipeline uploaded to Drive
    (a few MB — seconds, not minutes), keep a disk copy, and read out everything
    the UI needs. If Drive is unreachable, an existing local copy still serves.

    Returns the meta dict (+ df/path/generated_at), {"error": …} when there is
    nothing readable, or None when no store exists at all. Never raises: a
    corrupt / locked / version-mismatched file must degrade to the legacy live
    and upload modes, not paint a traceback over the whole app."""
    err = None
    fid = _store_folder_id()
    if fid:
        try:
            raw, _stamp = live_data.fetch_store(fid)
            if raw:
                _DISK_CACHE.mkdir(exist_ok=True)
                tmp = _STORE_PATH.with_suffix(".tmp")
                tmp.write_bytes(raw)
                try:
                    tmp.replace(_STORE_PATH)
                except OSError:
                    pass                     # another session holds it open — keep old copy
        except Exception as e:
            err = str(e)
    if not _STORE_PATH.exists():
        return {"error": err} if err else None
    try:
        import duckdb
        con = duckdb.connect(str(_STORE_PATH), read_only=True)
        try:
            out = {k: json.loads(v)
                   for k, v in con.execute("SELECT key, value FROM meta").fetchall()}
            out["df"] = con.execute("SELECT * FROM compute").df()
        finally:
            con.close()
    except Exception as e:
        # Unreadable store (mid-build write lock, interrupted build with no meta
        # table, truncated upload, newer duckdb storage format, …).
        return {"error": f"the prebuilt data file could not be read ({e})"}
    if "df" not in out or "generated_at" not in out:
        return {"error": "the prebuilt data file is incomplete (rebuild it)"}
    out["path"] = str(_STORE_PATH)
    out["drive_error"] = err
    return out


@st.cache_data(show_spinner=False, ttl=_STORE_TTL_SECONDS)
def _snapshot_list(nonce):
    """Past weeks available in archive/, newest first. Never raises: history is
    a bonus, and a Drive hiccup here must not take the live dashboard with it."""
    fid = _store_folder_id()
    if not fid:
        return []
    try:
        return live_data.list_store_snapshots(fid)
    except Exception:
        return []


@st.cache_data(show_spinner=False)
def _archived_marked(day):
    """That week's dated marked workbook in archive/, or None.

    None is a real answer, not an error: weeks refreshed before the archive
    existed have no marked copy, and saying so beats serving today's file under
    an old week's name."""
    fid = _store_folder_id()
    if not fid:
        return None
    try:
        import archive
        return live_data.find_archived(
            fid, archive.snapshot_name(archive.MARKED_LABEL,
                                       datetime.strptime(day, "%Y-%m-%d").date()))
    except Exception:
        return None


@st.cache_data(show_spinner="Opening that week's dashboard…")
def _load_snapshot(file_id, name):
    """Render an ARCHIVED week instead of the current one.

    Kept separate from _load_store rather than parameterised, for two reasons:
    a snapshot is immutable so it wants permanent caching rather than the live
    store's 30-minute TTL, and it must never be written to _STORE_PATH — doing
    so would leave last month's data sitting where the live loader expects
    today's, and the app would serve it as current long after the user moved on.
    """
    path = _DISK_CACHE / f"snap_{name}"
    try:
        if not path.exists():
            _DISK_CACHE.mkdir(exist_ok=True)
            tmp = path.with_suffix(".tmp")
            tmp.write_bytes(live_data.fetch_store_snapshot(file_id))
            tmp.replace(path)
        import duckdb
        con = duckdb.connect(str(path), read_only=True)
        try:
            out = {k: json.loads(v)
                   for k, v in con.execute("SELECT key, value FROM meta").fetchall()}
            out["df"] = con.execute("SELECT * FROM compute").df()
        finally:
            con.close()
    except Exception as e:
        return {"error": f"that week's snapshot could not be opened ({e})"}
    if "df" not in out or "generated_at" not in out:
        return {"error": "that week's snapshot is incomplete"}
    out["path"] = str(path)
    out["drive_error"] = None
    return out


@st.cache_data(show_spinner=False)
def _store_grid(path, batch, stamp):
    """One batch's roster grid from the store. `stamp` (generated_at_iso) keys
    the cache so a fresh store invalidates old grids. Returns None rather than
    raising if the store became unreadable since it was loaded."""
    try:
        import duckdb
        con = duckdb.connect(path, read_only=True)
        try:
            return con.execute(f'SELECT * FROM "grid_{batch}"').df()
        finally:
            con.close()
    except Exception:
        return None


@st.cache_data(show_spinner=False)
def _grid(roster_bytes, sheet_name):
    return dc.roster_grid(roster_bytes, sheet_name)


@st.cache_data(show_spinner=False)
def _sheet_map(roster_bytes):
    return dc.batch_sheet_map(roster_bytes)


@st.cache_data(show_spinner="Syncing from Google Drive & marking… (first run is the slow one)")
def _load_live_marked(nonce, mode):
    """Fetch from Drive AND mark, in one cached step that returns only the small
    results (marked roster bytes, report, warnings). The raw attendee CSVs are
    released as soon as marking is done instead of living in the cache forever —
    that retained 100+ MB was the main reason the app got slow / crashed.

    New sessions are only added on weekends, so there's deliberately NO time-based
    expiry — the result stays cached (shared across all viewers) until someone
    clicks '🔄 Refresh from Google', which bumps `nonce` and clears the cache.
    Marking is additionally disk-memoized, so even a server restart doesn't redo
    it unless the roster or the attendee file set actually changed."""
    data = live_data.load_live()
    roster_bytes, l2_bytes = data["roster_bytes"], data["l2_bytes"]
    attendee_files = data["attendee_files"] or []
    marked, report, warnings, mark_error = roster_bytes, [], [], None
    if attendee_files:
        # Key on the sheets' modifiedTime stamps + the attendee file set, NOT on
        # the exported bytes: Google's xlsx export differs run-to-run for an
        # unchanged sheet, which would make a byte-keyed cache never hit.
        key = (f"{data.get('roster_stamp', '')}|{data.get('l2_stamp', '')}|{mode}|"
               + "\n".join(sorted(n for n, _ in attendee_files))).encode("utf-8", "replace")
        try:
            marked, report, warnings = _disk_memo(
                "marked", key,
                lambda: ac.process_files(roster_bytes, l2_bytes, list(attendee_files),
                                         mode=mode, values_only=True))
        except Exception as e:
            mark_error = str(e)
    return {
        "marked_bytes": marked,
        "l2_bytes": l2_bytes,
        "report": report,
        "warnings": warnings,
        "source": data["source"],
        "n_attendee_files": len(attendee_files),
        "mark_error": mark_error,
    }


# ───────────────────────────── small helpers ─────────────────────────────────
def mask_email(e: str) -> str:
    e = (e or "").strip()
    if "@" not in e:
        return "•••" if e else ""
    local, dom = e.split("@", 1)
    keep = local[:3] if len(local) > 3 else local[:1]
    return f"{keep}…@{dom}"


def mask_phone(p: str) -> str:
    digits = "".join(ch for ch in str(p or "") if ch.isdigit())
    return ("•" * max(0, len(digits) - 4)) + digits[-4:] if digits else ""


# ───────────────────────────── header ────────────────────────────────────────
st.title("📊 Be10X — AI CAP Attendance")

if "nonce" not in st.session_state:
    st.session_state.nonce = 0


# ───────────────────────────── sidebar: source + controls ────────────────────
live_ready = live_data.config_present()
# Prefer the prebuilt store: with store_folder_id set it is the real source and
# refreshes from Drive. A local-only store (someone ran `pipeline.py --no-upload`)
# is still used — that's the point of building it — but it can never be refreshed
# from Drive, so say so out loud rather than serving stale data silently.
_store_configured = bool(_store_folder_id())
_store_local_only = not _store_configured and _STORE_PATH.exists()
_store_available = _store_configured or _store_local_only

with st.sidebar:
    st.header("① Data source")

    if st.button("🔄 Refresh from Google", width='stretch',
                 disabled=not (live_ready or _store_available),
                 help="Pull the latest prebuilt dashboard from Drive "
                      "(the pipeline rebuilds it every Monday 06:00 IST)"):
        st.session_state.nonce += 1
        st.cache_data.clear()
        st.rerun()

    st.header("② Count basis")
    basis = st.radio(
        "Who counts as batch strength?",
        ["Active only (exclude refunds)", "All enrolled"],
        index=0,
        help="Sets the default for the Roster tab’s ‘Active learners only’ filter. "
             "Dashboard percentages are always measured against total strength.",
    )

    # The matching rule only means something when THIS server does the marking.
    # With prebuilt data the marking already happened in the pipeline, so showing
    # a live-looking radio here would be a lie.
    mode = "exact"
    if not _store_available:
        with st.expander("⚙️ Matching rule (advanced)"):
            mode_label = st.radio(
                "How to match a student to an attendee",
                ["Exact — Registered mail + number (recommended)",
                 "Inclusive — also WhatsApp / broadcast, last-10-digit phone"],
                index=0, label_visibility="collapsed",
            )
            mode = "exact" if mode_label.startswith("Exact") else "inclusive"


# ───────────────────────────── acquire the data ──────────────────────────────
# Preferred: the prebuilt store (instant). Fallbacks: legacy fetch+mark here,
# then manual upload. Only the store path is exercised once the pipeline runs.
marked_bytes = l2_bytes = None
report, warnings = [], []
n_attendee_files = 0
upload_attendee_names = ()
source_label = None
live_failed = False
mark_error = None
store = None
store_mode = False

# ── which week are we looking at? ────────────────────────────────────────────
# Every Monday the pipeline archives the store it just built, so each past week's
# dashboard survives verbatim rather than being overwritten. Offer them here, or
# the archive is data nobody can actually reach.
_snapshots = _snapshot_list(st.session_state.nonce) if _store_configured else []
_viewing = None
if _snapshots:
    with st.sidebar:
        _labels = ["Latest (live)"] + [
            datetime.strptime(s["date"], "%Y-%m-%d").strftime("%d %b %Y")
            for s in _snapshots]
        _pick_week = st.selectbox(
            "Week", _labels, index=0, key="week_pick",
            help="Past weeks are archived every Monday. Choosing one shows the "
                 "dashboard exactly as it stood then — attendance, day-1, "
                 "forecast and roster all as they were.")
        if _pick_week != "Latest (live)":
            _viewing = _snapshots[_labels.index(_pick_week) - 1]

if _viewing:
    loaded = _load_snapshot(_viewing["id"], _viewing["name"])
    if loaded and "df" in loaded:
        store = loaded
        store_mode = True
        report, warnings = store["report"], store["warnings"]
        source_label = store["source"]
        # Say it on the PAGE, not just the sidebar. Someone screenshotting a
        # number from a historical week must not be able to mistake it for today.
        st.warning(
            f"📅 **Viewing the archived week of "
            f"{datetime.strptime(_viewing['date'], '%Y-%m-%d'):%d %b %Y}** — this "
            f"is the dashboard as it stood then (built {store['generated_at']}), "
            "not current data. Switch **Week** back to “Latest (live)” in the "
            "sidebar for today's numbers.")
    else:
        st.sidebar.error(
            "Couldn’t open that week: "
            + str((loaded or {}).get("error", "unknown")))
        _viewing = None

if not _viewing and _store_available:
    loaded = _load_store(st.session_state.nonce)
    if loaded and "df" in loaded:
        store = loaded
        store_mode = True
        report, warnings = store["report"], store["warnings"]
        source_label = store["source"]
        if store.get("drive_error"):
            st.sidebar.warning("Drive unreachable — showing the last downloaded "
                               f"data.\n\n{store['drive_error']}")
        if _store_local_only:
            st.sidebar.warning(
                f"Using a **locally built** store from {store['generated_at']}. "
                "🔄 Refresh cannot update it — no `store_folder_id` is configured. "
                "Re-run `pipeline.py`, or delete `.cache/attendance.duckdb` to read "
                "Google Drive directly."
            )
        else:
            st.sidebar.caption(f"📅 Data as of **{store['generated_at']}** · "
                               "rebuilt every Monday 06:00 IST")
        pipeline_mode = (store.get("stamps") or {}).get("mode") or "exact"
        st.sidebar.caption(f"Matching rule: **{pipeline_mode}** (set by the pipeline)")
    elif loaded and loaded.get("error"):
        st.sidebar.error(f"Couldn’t load the prebuilt dashboard:\n\n{loaded['error']}"
                         + ("\n\nFalling back to reading Google Drive directly."
                            if live_ready else ""))

if not store_mode and live_ready:
    try:
        live = _load_live_marked(st.session_state.nonce, mode)
        marked_bytes = live["marked_bytes"]
        l2_bytes = live["l2_bytes"]
        report, warnings = live["report"], live["warnings"]
        source_label = live["source"]
        n_attendee_files = live["n_attendee_files"]
        mark_error = live["mark_error"]
    except Exception as e:  # fall back to uploads, but tell the user why
        live_failed = True
        st.sidebar.error(f"Couldn’t read from Google Drive:\n\n{e}")

if not store_mode and marked_bytes is None:
    with st.sidebar:
        if not live_ready:
            st.caption("Google not connected yet — upload files below. "
                       "See **SETUP_LIVE.md** to make it automatic.")
        up_roster = st.file_uploader("Master Roster (.xlsx)", type=["xlsx"])
        with st.expander("Optional: mark fresh attendance"):
            up_l2 = st.file_uploader("L2 schedule (.xlsx)", type=["xlsx"])
            up_zip = st.file_uploader("Zoom attendee reports (.zip)", type=["zip"])
    if up_roster is None:
        st.info("👈 **Upload your Master Roster** in the sidebar to begin "
                "(or set up Google for automatic updates — see SETUP_LIVE.md).")
        st.stop()
    roster_bytes = up_roster.getvalue()
    l2_bytes = up_l2.getvalue() if up_l2 else None
    marked_bytes = roster_bytes
    if up_zip is not None:
        import zipfile
        with zipfile.ZipFile(io.BytesIO(up_zip.getvalue())) as z:
            attendee_files = [(n, z.read(n)) for n in z.namelist() if not n.endswith("/")]
        n_attendee_files = len(attendee_files)
        upload_attendee_names = tuple(n for n, _ in attendee_files)
        source_label = f"Manual upload — roster + {n_attendee_files} attendee files"
        try:
            marked_bytes, report, warnings = _mark(roster_bytes, l2_bytes, tuple(attendee_files), mode)
        except Exception as e:
            mark_error = str(e)
    else:
        source_label = "Manual upload — roster only"

if mark_error:
    st.error(f"Marking failed, showing the roster as-is: {mark_error}")

if store_mode:
    df = store["df"]
else:
    try:
        df = _compute(marked_bytes)
    except Exception as e:
        st.error(f"Couldn’t read the roster: {e}")
        st.stop()

if df.empty:
    st.warning("No batch sheets with session columns were found.")
    st.stop()

marked = df[df["HasData"]].copy()
if marked.empty:
    st.warning("No sessions have attendance marked yet in this roster.")
    st.stop()

# basis-driven column names
active_mode = basis.startswith("Active")
PCT = "PctActive" if active_mode else "PctAll"
PRES = "PresentActive" if active_mode else "PresentAll"
DEN = "Active" if active_mode else "Total"
den_label = "active learners" if active_mode else "enrolled"

# status line
if store_mode:
    auto = "🟢 Prebuilt data"
    st.caption(f"{auto} · {source_label} · data as of {store['generated_at']}")
else:
    auto = "🟢 Live from Google Drive" if (live_ready and not live_failed) else "📤 Manual upload"
    st.caption(f"{auto} · {source_label} · refreshed {datetime.now():%d %b %Y, %H:%M}")
if report:
    new_n = sum(1 for r in report if r["kind"] == "NEW")
    st.success(f"✅ Auto-marked attendance: {len(report)} session column(s) across "
               f"{len({r['batch'] for r in report})} batch(es) "
               f"({new_n} new, {len(report)-new_n} re-marked).")


def _marked_roster_download(key: str) -> None:
    """The full marked workbook. In store mode the ~8 MB xlsx isn't in memory —
    it's fetched from Drive only when someone actually asks for it."""
    if not store_mode:
        st.download_button(
            "⬇️ Download marked roster — ALL sessions (.xlsx)",
            data=marked_bytes,
            file_name="Master_Batch_Rosters_marked.xlsx",
            mime=XLSX_MIME, type="primary", key=key,
        )
        return
    # While viewing an archived week, `marked_xlsx_file_id` in that old store
    # points at Master_Batch_Rosters_marked.xlsx, which the pipeline REPLACES
    # every Monday — following it would hand over TODAY's workbook labelled as
    # that week's. Use the dated copy in archive/ instead, and say plainly when
    # a week predates archiving rather than serving the wrong file.
    if _viewing:
        _arch = _archived_marked(_viewing["date"])
        if not _arch:
            st.caption(
                f"No marked workbook was archived for {_viewing['date']} — that "
                "week ran before the archive existed. Only weeks refreshed since "
                "then have one.")
            return
        fid, token = _arch["id"], f"{_arch['id']}@{_viewing['date']}"
        fname = _arch["name"]
    else:
        fid = store.get("marked_xlsx_file_id") or ""
        token = f"{fid}@{store['generated_at_iso']}" if fid else ""
        fname = "Master_Batch_Rosters_marked.xlsx"
    if not fid:
        st.caption("The marked workbook isn’t on Drive yet — it is uploaded by a "
                   "full pipeline run (not by `--no-upload`).")
        return
    # For the LIVE file the Drive id never changes (the pipeline replaces it in
    # place), so the token folds in the build time — otherwise a long-lived
    # session keeps serving last week's workbook after a refresh.
    if st.session_state.get("marked_xlsx_token") != token:
        if st.button("⬇️ Prepare marked roster download (.xlsx)", key=f"prep_{key}",
                     help="Fetches the full marked workbook (~8 MB) from Drive"):
            try:
                with st.spinner("Fetching the marked workbook from Drive…"):
                    st.session_state["marked_xlsx"] = live_data.fetch_file_bytes(
                        live_data._drive_service(), fid)
                    st.session_state["marked_xlsx_token"] = token
            except Exception as e:
                st.error(f"Couldn’t fetch the marked workbook from Drive: {e}")
            else:
                st.rerun()
    if st.session_state.get("marked_xlsx_token") == token:
        st.download_button(
            "⬇️ Download marked roster — ALL sessions (.xlsx)",
            data=st.session_state["marked_xlsx"],
            file_name=fname,
            mime=XLSX_MIME, type="primary", key=key,
        )


# Prominent: the full marked roster (all sessions) — same deliverable as the original marker
_marked_roster_download("dl_top")
st.caption("The complete marked workbook — every batch, all sessions, Present/Absent.")


# Batch list for the Roster tab (the new Dashboard has its own selector).
all_batches = sorted(marked["Batch"].unique(), key=dc.batch_key)


# ───────────────────────────── tabs ──────────────────────────────────────────
(tab_dash, tab_sessions, tab_roster, tab_day1, tab_fcst, tab_recap,
 tab_trainer, tab_students, tab_bsiai) = st.tabs(
    ["📊 Dashboard", "📚 Sessions", "📋 Roster (marked attendance)",
     "🎯 Day-1 analysis", "🔮 Forecast", "🏆 Recap", "🎓 Trainers",
     "🧑‍🎓 Students", "🧩 BSIAI"]
)

# ============================ TAB 1 — DASHBOARD ==============================
with tab_dash:
    if store_mode:
        # Everything was prebuilt by pipeline.py — render straight from the store.
        DATA, summary = store["DATA"], store["summary"]
        note = f"🟢 Data as of {store['generated_at']} · rebuilt every Monday 06:00 IST"
    else:
        # Legacy: build from the roster this server just marked.
        if live_ready and not live_failed:
            names = _attendee_names(st.session_state.nonce)
            note = f"🟢 Live — showing the updated roster · refreshed {datetime.now():%d %b %Y, %H:%M}"
        else:
            names = upload_attendee_names
            note = "📤 Showing the uploaded roster"
        DATA, summary = _build_dashboard(marked_bytes, names, l2_bytes)
    dash_view.render(DATA, summary, note)

# ============================ TAB 2 — ROSTER =================================
with tab_roster:
    st.caption("The marked attendance, student-by-student, exactly like the roster sheet.")
    if store_mode:
        grid_batches = [b for b in store["batches"] if b in set(all_batches)] or store["batches"]
    else:
        smap = _sheet_map(marked_bytes)
        grid_batches = [b for b in all_batches if b in smap]
    ctop1, ctop2 = st.columns([3, 2])
    pick = ctop1.selectbox("Batch", grid_batches,
                           index=0 if grid_batches else None, key="roster_pick")
    show_pii = ctop2.checkbox("Show full contact details (real PII)", value=False,
                              help="Off by default — emails/phones are masked.")
    g = None
    if pick:
        if store_mode:
            g = _store_grid(store["path"], pick, store["generated_at_iso"])
            g = g.copy() if g is not None else None
        else:
            g = _grid(marked_bytes, smap[pick]).copy()
    if g is None:
        st.info("No roster rows to show for this batch — try 🔄 Refresh from Google."
                if pick else "No batch sheets to show.")
    else:
        active_only = st.checkbox("Active learners only", value=active_mode, key="roster_active")
        if active_only:
            g = g[g["Active"].astype(bool)]
        sess_cols = [c for c in g.columns if c not in ("Email", "Phone", "Active", "Present")]
        n_students = len(g)
        n_present_any = int((g["Present"] > 0).sum())
        m1, m2, m3 = st.columns(3)
        m1.metric("Students shown", f"{n_students:,}")
        m2.metric("Attended ≥1 session", f"{n_present_any:,}")
        m3.metric("Sessions", len(sess_cols))

        disp = g.copy()
        if not show_pii:
            disp["Email"] = disp["Email"].map(mask_email)
            disp["Phone"] = disp["Phone"].map(mask_phone)
        disp = disp[["Email", "Phone", "Active", "Present"] + sess_cols]

        def _hl(v):
            s = str(v).lower()
            if s == "present":
                return "background-color: #d8f3dc; color:#1b4332"
            if s == "absent":
                return "background-color: #ffe5e5; color:#7d1128"
            return ""
        styled = disp.style.map(_hl, subset=sess_cols)  # .map (pandas >=2.1; applymap removed in 3.0)
        st.dataframe(styled, width='stretch', hide_index=True, height=520)

    st.divider()
    _marked_roster_download("dl_roster")

# ========================= TAB 3 — DAY-1 ANALYSIS ============================
# Day-one (and latest-session) attendance cut against payment (col I) and close
# type (col J), for the newest batches. Built by pipeline.py — day one is the
# earliest session the L2 schedule registers as a real class, so marketing
# walkthroughs never get mistaken for class 1 — and carried in the store as
# aggregates only (no PII). New batches appear on their own each Monday.
with tab_day1:
    import streamlit.components.v1 as _components
    _day1 = (store or {}).get("day1") if store_mode else None
    if not store_mode:
        st.info("This tab is built by the weekly pipeline. It appears once the app "
                "is reading the prebuilt store — see **SETUP_PIPELINE.md**.")
    elif not _day1 or not _day1.get("batches"):
        st.warning("No day-1 analysis in this build of the data.")
        for _s in (_day1 or {}).get("skipped", []):
            st.caption("• " + _s)
    else:
        _all = _day1["batches"]
        _codes = [b["batch"] for b in _all]
        # Batch choice lives HERE, not inside the page: st.components.html ignores
        # a frame-height message when a height is given, so the only way to avoid
        # either a nested scrollbar or a screen of dead space is for Python to
        # know how many batches it is about to draw.
        _pick = st.multiselect("Batches", _codes, default=_codes, key="day1_batches",
                               help="The newest batches; the window rolls on its own "
                                    "as new ones start.")
        _sel = [b for b in _all if b["batch"] in _pick] or _all
        _payload = dict(_day1, batches=_sel)
        _payload["generated"] = store["generated_at"]
        _payload["sheet_url"] = (
            "https://docs.google.com/spreadsheets/d/"
            f"{st.secrets.get('drive', {}).get('roster_id', '')}/edit"
            if _store_configured or live_ready else "")
        _tpl = (Path(__file__).parent / "day1_template.html").read_text(encoding="utf-8")
        # `</` must be escaped or a session title containing "</script>" would
        # terminate the inline script and blank the tab with no error shown.
        _blob = json.dumps(_payload).replace("</", "<\\/")
        # Measured: ~3.3k px of fixed chrome + ~1.15k per batch of chart rows.
        _components.html(_tpl.replace("__DATA__", _blob),
                         height=3350 + 1150 * len(_sel), scrolling=True)
# ========================== TAB 4 — FORECAST =================================
# Predicted attendance for sessions that have not run yet, built by pipeline.py
# from the dashboard's own DATA plus the Master Curriculum Schedule. Aggregates
# only. Everything shown here is a PREDICTION — the copy says so in every place
# a number could otherwise be mistaken for a measurement.
with tab_fcst:
    _f = (store or {}).get("forecast") if store_mode else None

    if not store_mode:
        st.info("The Forecast tab reads the prebuilt store. It is empty in "
                "upload / legacy-live mode because the curriculum schedule is a "
                "separate Sheet the app does not fetch at runtime.")
    elif _f is None:
        st.warning(
            "**The forecast is not configured yet.** Add `CURRICULUM_ID` (repo "
            "secret, or `curriculum_id` under `[drive]` in secrets.toml) pointing "
            "at the Master Curriculum Schedule, and share that Sheet with the "
            "service account. The next refresh will fill this tab in."
        )
    elif not _f.get("sessions"):
        st.warning("The forecast produced no sessions in the last refresh.")
        for _w in _f.get("warnings", []):
            st.caption("• " + str(_w))
    else:
        import pandas as _pd

        _acc = _f.get("accuracy") or {}
        _rows = _f["sessions"]

        st.caption(
            f"Predicted attendance for the next {_f['horizon_weeks']} weeks, from "
            f"the batch's own attendance so far and the decay curve measured "
            f"across every past batch. Built {store['generated_at']}."
        )

        m1, m2, m3, m4 = st.columns(4)
        # Count SESSIONS, not per-batch rows: one session that three batches sit
        # in is one thing to run and staff, not three. The rows are still there
        # in the detail table and the full CSV.
        m1.metric("Sessions ahead", f"{len(_f.get('by_session') or []):,}",
                  help=f"{len(_rows):,} batch-slots across them — a session that "
                       "several batches attend counts once here.")
        m2.metric("Predicted attendees", f"{sum(r['pred'] for r in _rows):,}")
        if _acc.get("mape_1_4wk") is not None:
            m3.metric("Typical error, 1–4 wks", f"±{_acc['mape_1_4wk']}%",
                      help="Mean absolute % error on headcount, measured by "
                           "replaying every past batch: fit on what was known at "
                           "the time, score against what actually happened.")
        if _acc.get("baseline_mape_1_4wk") is not None:
            m4.metric("Naive baseline", f"±{_acc['baseline_mape_1_4wk']}%",
                      help="What you would get by assuming the next session "
                           "repeats the last one's attendance rate. The model is "
                           "only worth having while it beats this.")

        for _w in _f.get("warnings", []):
            st.warning(str(_w))

        # ── every session, batch-wise and total ──────────────────────────────
        # One row per SESSION with a column per batch, because that is the unit
        # the programme is staffed and briefed in: "Build Your Marketing OS on
        # 6 Sep draws ~230 across B35/B36/B37", not three unrelated numbers.
        st.subheader("Every session — batch-wise and total")
        _sess = _f.get("by_session") or []
        _all_b = sorted({b for s in _sess for b in s["batches"]}, key=dc.batch_key)

        # All three filters are multiselects: they are type-to-search and apply on
        # selection. A free-text box would only commit on Enter, which reads as a
        # broken filter to anyone who types and then looks straight at the table.
        c1, c2, c3 = st.columns([2, 2, 3])
        _fdate = c1.multiselect("Date", list(dict.fromkeys(
            _pd.to_datetime([s["date"] for s in _sess]).strftime("%a %d %b"))),
            key="fcst_date")
        _fpod = c2.multiselect("Pod", sorted({s["pod"] for s in _sess}),
                               key="fcst_pod")
        _ftopic = c3.multiselect("Topic — type to search",
                                 sorted({s["topic"] for s in _sess}),
                                 key="fcst_topic")

        _v = _sess
        if _fpod:
            _v = [s for s in _v if s["pod"] in _fpod]
        if _ftopic:
            _v = [s for s in _v if s["topic"] in _ftopic]
        _tbl = []
        for s in _v:
            _lbl = _pd.to_datetime(s["date"]).strftime("%a %d %b")
            if _fdate and _lbl not in _fdate:
                continue
            _r = {"Date": _lbl, "Pod": s["pod"], "Topic": s["topic"],
                  "Trainer": s["trainer"] or "—"}
            for b in _all_b:
                _pb = s["per_batch"].get(b)
                _r[b] = _pb["pred"] if _pb else None
            _r["TOTAL"] = s["pred"]
            _r["Range"] = f"{s['lo']:,} – {s['hi']:,}"
            _r["Of"] = s["denom"]
            _r["%"] = s["pred_pct"]
            _r["≈"] = "≈" if s["any_assumed"] else ""
            _tbl.append(_r)

        if not _tbl:
            st.info("No sessions match those filters.")
        else:
            _tdf = _pd.DataFrame(_tbl)
            st.dataframe(
                _tdf, width='stretch', hide_index=True, height=520,
                column_config={
                    **{b: st.column_config.NumberColumn(b, format="%d", width="small")
                       for b in _all_b},
                    "TOTAL": st.column_config.NumberColumn(
                        "TOTAL", format="%d",
                        help="Sum across the batches running this session"),
                    "Of": st.column_config.NumberColumn(
                        "Of", format="%d",
                        help="Combined size of the groups invited"),
                    "%": st.column_config.NumberColumn("%", format="%.1f%%"),
                    "≈": st.column_config.TextColumn(
                        "≈", width="small",
                        help="At least one batch here has not started, so its "
                             "size is assumed rather than measured"),
                })
            st.caption(
                f"Showing **{len(_tbl)}** of {len(_sess)} sessions · "
                f"**{sum(r['TOTAL'] for r in _tbl):,}** predicted attendees. "
                "Blank cell = that batch is not scheduled for this session. "
                "The range widens for dates further out, and totals add each "
                "batch's range rather than assuming their errors cancel — "
                "batches sharing a date also share their bad days.")

        _c1, _c2 = st.columns(2)
        _c1.download_button(
            "⬇️ Sessions, batch-wise (.csv)",
            data=(_pd.DataFrame(_tbl).to_csv(index=False).encode("utf-8")
                  if _tbl else b""),
            file_name=f"forecast_by_session_{_f.get('generated_for','')}.csv",
            mime="text/csv", key="dl_fcst_sess", disabled=not _tbl)
        _c2.download_button(
            "⬇️ Full detail, one row per batch (.csv)",
            data=_pd.DataFrame(_rows).to_csv(index=False).encode("utf-8"),
            file_name=f"attendance_forecast_{_f.get('generated_for','')}.csv",
            mime="text/csv", key="dl_fcst")

        with st.expander("How this is calculated, and how much to trust it"):
            st.markdown(
                "**predicted = group size × decay curve(week) × batch offset × "
                "pod multiplier**\n\n"
                "- **Decay curve** — attendance against weeks-since-batch-start, "
                "pooled over every past batch. The shape is very stable: batches "
                "open near 50% and fall about 10% a week.\n"
                "- **Batch offset** — how this batch is actually tracking against "
                "that curve. A batch running cold shows up here first.\n"
                "- **Pod multiplier** — how a pod draws relative to its own "
                "batch on the same day. **This is the weak layer**: the pod split "
                "only began in late Aug 2026, so it rests on few sessions and is "
                "shrunk toward the batch average.\n\n"
                "Accuracy is re-measured on live data every refresh by replaying "
                "history, so if the model stops working this page will say so.")
            _cA, _cB = st.columns(2)
            with _cA:
                st.caption("**Error by weeks ahead** (measured)")
                _mh = _acc.get("mape_by_horizon") or {}
                if _mh:
                    st.dataframe(_pd.DataFrame(
                        {"Weeks ahead": list(_mh), "Error ±%": list(_mh.values())}),
                        width='stretch', hide_index=True, height=240)
            with _cB:
                st.caption("**Pod multipliers** (>1 attends more than its batch)")
                _pmd = _f.get("pod_multiplier") or {}
                if _pmd:
                    st.dataframe(_pd.DataFrame(
                        [{"Pod": k, "×": v["mult"], "Sessions seen": v["n"]}
                         for k, v in sorted(_pmd.items(),
                                            key=lambda kv: -kv[1]["mult"])]),
                        width='stretch', hide_index=True, height=240)
            st.caption("**Batch offset** — ×1.00 is exactly on the curve; "
                       "below 1 means that batch is running cold")
            _bo = _f.get("batch_offset") or {}
            if _bo:
                st.dataframe(_pd.DataFrame(
                    [{"Batch": k, "×": v} for k, v in sorted(
                        _bo.items(), key=lambda kv: dc.batch_key(kv[0]))]),
                    width='stretch', hide_index=True, height=240)

# ============================ TAB 2 - SESSIONS ===============================
with tab_sessions:
    _S = (store or {}).get("sessions") if store_mode else None
    if not store_mode:
        st.info("The session browser reads the prebuilt data file.")
    elif not _S:
        st.warning("No sessions in the last refresh.")
    else:
        import pandas as _pd
        st.caption(
            "Every logged session across every batch. Select a row for the full "
            "breakdown. **Attendance % = present ÷ batch strength**, taken from "
            "the roster automatically — so it is filled in for all "
            f"{len({s['batch'] for s in _S})} batches, not only the ones somebody "
            "remembered to configure."
        )

        # ── filters ──────────────────────────────────────────────────────────
        _dates = sorted({s["date"] for s in _S})
        f1, f2, f3 = st.columns([2, 1, 1])
        _dr = f1.date_input(
            "Date range",
            value=(_dt.date.fromisoformat(_dates[0]), _dt.date.fromisoformat(_dates[-1])),
            min_value=_dt.date.fromisoformat(_dates[0]),
            max_value=_dt.date.fromisoformat(_dates[-1]), key="ss_dates")
        _fb = f2.multiselect("Batch", sorted({s["batch"] for s in _S},
                                             key=dc.batch_key), key="ss_batch")
        _fp = f3.multiselect("POD", sorted({s["pod"] for s in _S if s["pod"]}),
                             key="ss_pod")
        g1, g2, g3 = st.columns([2, 1, 2])
        _ft = g1.multiselect("Trainer", sorted({t for s in _S
                                                for t in (s.get("trainers") or [])}),
                             key="ss_trainer")
        _fy = g2.multiselect("Trainer type",
                             sorted({s["trainer_type"] for s in _S
                                     if s.get("trainer_type")}), key="ss_ttype")
        _fq = g3.text_input("Search title", key="ss_q").strip().lower()

        _v = _S
        if isinstance(_dr, (tuple, list)) and len(_dr) == 2:
            _a, _b = _dr[0].isoformat(), _dr[1].isoformat()
            _v = [s for s in _v if _a <= s["date"] <= _b]
        if _fb:
            _v = [s for s in _v if s["batch"] in _fb]
        if _fp:
            _v = [s for s in _v if s["pod"] in _fp]
        if _ft:
            _v = [s for s in _v if set(s.get("trainers") or []) & set(_ft)]
        if _fy:
            _v = [s for s in _v if s.get("trainer_type") in _fy]
        if _fq:
            _v = [s for s in _v if _fq in (s["topic"] or "").lower()]

        # ── headline numbers, over the FILTERED set ──────────────────────────
        _rated = [s for s in _v if s.get("rating") is not None]
        _rn = sum(s["rating_n"] for s in _rated)
        _merged = _polls_mod.merge_dists(s.get("dist") for s in _v)
        _pres = sum(s["present"] for s in _v)
        _inv = sum(s["total"] for s in _v)
        m1, m2, m3, m4, m5 = st.columns(5)
        m1.metric("Sessions", f"{len(_v):,}")
        m2.metric("Attendance", f"{_pres / _inv * 100:.1f}%" if _inv else "—",
                  help=f"{_pres:,} present of {_inv:,} invited, pooled — not a "
                       "mean of per-session percentages.")
        m3.metric("Avg overall",
                  f"{sum(s['rating'] * s['rating_n'] for s in _rated) / _rn:.2f}"
                  if _rn else "—",
                  help="Weighted by responses, so a 12-response session does not "
                       "outweigh a 900-response one.")
        _trd = [s for s in _v if s.get("rating_trainer") is not None]
        _trn = sum(s["rating_n"] for s in _trd)
        m4.metric("Avg trainer",
                  f"{sum(s['rating_trainer'] * s['rating_n'] for s in _trd) / _trn:.2f}"
                  if _trn else "—")
        _nps = _polls_mod.nps_from_dist(_merged.get("recommend"))
        m5.metric("NPS", f"{_nps:+d}" if _nps is not None else "—",
                  help="Promoter 5, passive 4, detractor 1-3, computed from the "
                       "SUMMED 1-5 histograms of every session shown — not an "
                       "average of their NPS percentages.")

        # ── the table ────────────────────────────────────────────────────────
        _tbl = _pd.DataFrame([{
            "Date": s["date"], "Trainer": s.get("trainer") or "—",
            "Type": s.get("trainer_type") or "—",
            "Title": s["topic"], "Batch": s.get("l2_batch") or s["batch"],
            "POD": s["pod"] or "—",
            "Attendance %": s["pct"], "Present": s["present"],
            "Invited": s["total"], "vs curve": s.get("index"),
            "Overall": s.get("rating"), "Trainer rating": s.get("rating_trainer"),
            "NPS": s.get("nps"), "Responses": s["rating_n"],
            "Simulive": s.get("session_type") or "",
        } for s in _v])
        _sel = st.dataframe(
            _tbl, width='stretch', hide_index=True, height=460,
            on_select="rerun", selection_mode="single-row",
            column_config={
                "Overall": st.column_config.ProgressColumn(
                    "Overall", min_value=0, max_value=5, format="%.2f"),
                "Attendance %": st.column_config.NumberColumn(format="%.1f%%"),
                "vs curve": st.column_config.NumberColumn(
                    format="%.2fx", help="Above 1.00 beats what a cohort of this "
                                         "age normally draws."),
            })

        st.caption(
            "**Duration** and **Peak** are deliberately absent. Both live only in "
            "the Zoom attendee report's own header, which the weekly job does not "
            "fetch for past sessions — and Zoom's 'Actual Duration' tracks when "
            "the host finally left, not when teaching ended (measured: 112 to 279 "
            "minutes for nominally 3-hour classes). **Simulive** is blank unless "
            "L2 says so; only 49 of 1,610 mentor rows record it, so a blank means "
            "not recorded rather than Live."
        )

        _rowsel = (_sel.selection.rows if getattr(_sel, "selection", None) else [])
        if _rowsel:
            s = _v[_rowsel[0]]
            st.divider()
            st.subheader(s["topic"] or "Session")
            st.caption(f"{s.get('l2_batch') or s['batch']} · {s['date_lbl']} "
                       f"({s['date']})"
                       + (f" · {s['pod']}" if s["pod"] else "")
                       + (f" · {s['trainer']}" if s.get("trainer") else "")
                       + (f" ({s['trainer_type']})" if s.get("trainer_type") else ""))
            d1, d2, d3, d4 = st.columns(4)
            d1.metric("Present", f"{s['present']:,}")
            d2.metric("Invited", f"{s['total']:,}")
            d3.metric("Attendance", f"{s['pct']:.1f}%" if s["pct"] else "—")
            d4.metric("vs curve", f"{s['index']:.2f}x" if s.get("index") else "—",
                      help=(f"the curve expects {s['expected_pct']:.1f}% for a "
                            f"{s['batch']} session {s['wk']} weeks in"
                            if s.get("expected_pct") else
                            "no curve point for this week — not scored"))
            _dist = (s.get("dist") or {})
            if any(_dist.values()):
                st.markdown("**How the room rated it**")
                _cols = st.columns(3)
                for _c, _kind in zip(_cols, ("session", "trainer", "recommend")):
                    _h = _dist.get(_kind) or {}
                    if not sum(_h.values()):
                        continue
                    with _c:
                        st.caption(_kind.title())
                        st.bar_chart(_pd.DataFrame(
                            {"count": [_h.get(str(i), 0) for i in range(1, 6)]},
                            index=[str(i) for i in range(1, 6)]))
            else:
                st.caption("No poll was run for this session.")

# ============================ TAB 5 - RECAP ==================================
with tab_recap:
    _r = (store or {}).get("recap") if store_mode else None
    if not store_mode:
        st.info("The recap is built by the weekly pipeline, so it needs the "
                "prebuilt data file. This server is running in live/upload mode.")
    elif not _r or not _r.get("weeks"):
        st.warning("No recap in the last refresh."
                   + ("  \n• " + "  \n• ".join(_r.get("warnings") or []) if _r else ""))
    else:
        import pandas as _pd
        _lat = _r["latest"]
        st.caption(
            "Every headline here is a **residual** — what a session actually drew "
            "over what the decay curve says a batch of that age, level and pod "
            "should draw. Ranking on raw attendance just crowns the youngest "
            "cohort every week, because attendance falls about 10% a week over a "
            f"batch's life. An index of 1.00 is exactly on curve. "
            f"Built {store['generated_at']}."
        )

        def _d(v, pct=False, signed=True):
            if v is None:
                return None
            return f"{v:+.1f}{'%' if pct else ''}" if signed else f"{v:.1f}"

        k1, k2, k3, k4, k5 = st.columns(5)
        k1.metric("Sessions", f"{_lat['sessions']:,}")
        k2.metric("Learners present", f"{_lat['present']:,}",
                  _d(_lat["delta"]["present"]))
        k3.metric("Attendance", f"{_lat['pct']:.1f}%" if _lat["pct"] else "—",
                  _d(_lat["delta"]["pct"], pct=True))
        k4.metric("vs curve", f"{_lat['index']:.2f}x" if _lat["index"] else "—",
                  _d(_lat["delta"]["index"]),
                  help="Above 1.00 means these sessions beat what cohorts of "
                       "their age normally draw. This is the one to watch — the "
                       "raw percentage falls every week by design.")
        k5.metric("NPS", f"{_lat['nps']:+d}" if _lat["nps"] is not None else "—",
                  _d(_lat["delta"]["nps"]),
                  help="Promoter 5, passive 4, detractor 1-3, on the poll's "
                       "recommend question.")

        if _r.get("awards"):
            st.subheader("This week")
            for _a, _col in zip(_r["awards"], st.columns(len(_r["awards"]))):
                with _col:
                    st.markdown(
                        f"**{_a['award']}**  \n"
                        f"### {_a['value']}  \n"
                        f"{_a['batch']} · {_a['topic'][:44]}"
                        + (f" · {_a['pod']}" if _a.get("pod") else "")
                        + (f"  \n_{_a['mentor']}_" if _a.get("mentor") else "")
                        + f"  \n<span style='color:#7a8598;font-size:12px'>{_a['why']}</span>",
                        unsafe_allow_html=True)

        st.subheader("Week by week")
        st.dataframe(_pd.DataFrame([{
            "Week of": w["week"], "Sessions": w["sessions"],
            "Batches": len(w["batches"]), "Present": w["present"],
            "Invited": w["invited"], "Attendance %": w["pct"],
            "vs curve": w["index"], "Rating": w["rating"], "NPS": w["nps"],
        } for w in reversed(_r["weeks"])]), width='stretch', hide_index=True)

        if _r.get("leaderboard"):
            st.subheader("Trainers, this week")
            st.caption("Ranked on the residual, so a trainer who taught an old "
                       "cohort is not punished for its age.")
            st.dataframe(_pd.DataFrame([{
                "Trainer": b["mentor"], "Sessions": b["sessions"],
                "Present": b["present"], "Attendance %": b["pct"],
                "vs curve": b["index"], "Rating": b["rating"], "NPS": b["nps"],
            } for b in _r["leaderboard"]]), width='stretch', hide_index=True)

# ============================ TAB 6 - TRAINERS ===============================
with tab_trainer:
    _t = (store or {}).get("trainers") if store_mode else None
    if not store_mode:
        st.info("Trainer rollups are built by the weekly pipeline.")
    elif not _t or not _t.get("trainers"):
        st.warning("No trainer data in the last refresh."
                   + ("  \n• " + "  \n• ".join(_t.get("warnings") or []) if _t else ""))
    else:
        import pandas as _pd
        st.caption(
            f"**{_t['n_raw']} spellings in L2 resolve to {_t['n_people']} people.** "
            "The Mentor cell is hand-typed, so one person arrives as 'Swapnil', "
            "'Swapnil Narayan' and 'Swapnil (Play Simulive)'. Where L2 carries an "
            "email it is trusted; otherwise a short name joins a longer one only "
            "when it is an unambiguous prefix of it."
        )
        _min = st.slider("Minimum sessions", 1, 25, 5, key="tr_min",
                         help="A trainer's index over one session is noise.")
        _rows = [x for x in _t["trainers"] if x["sessions"] >= _min]
        st.dataframe(_pd.DataFrame([{
            "Trainer": x["trainer"], "Sessions": x["sessions"],
            "Co-taught": x["co_taught"], "Batches": len(x["batches"]),
            "Present": x["present"], "Invited": x["invited"],
            "Attendance %": x["pct"], "vs curve": x["index"],
            "Rating": x["rating"], "Responses": x["rating_n"], "NPS": x["nps"],
            "First": x["first"], "Last": x["last"],
        } for x in _rows]), width='stretch', hide_index=True, height=520)
        st.caption(f"{len(_rows)} of {_t['n_people']} shown. **Co-taught** counts "
                   "sessions credited to more than one trainer — both get the "
                   "session, so totals across trainers exceed the session count.")

        if _t.get("merged"):
            with st.expander(f"Spellings merged ({len(_t['merged'])} people)"):
                st.dataframe(_pd.DataFrame(
                    [{"Shown as": k, "Merged from": ", ".join(v)}
                     for k, v in sorted(_t["merged"].items())]),
                    width='stretch', hide_index=True)
        if _t.get("ambiguous"):
            st.warning(
                "Left unmerged because the name could be more than one person: "
                + ", ".join(f"**{n}**" for n in _t["ambiguous"])
                + ". Add their email to L2's *Mentor's email* column to resolve them.")

# ============================ TAB 7 - STUDENTS ===============================
with tab_students:
    if not store_mode:
        st.info("The per-student view reads the prebuilt roster grids.")
    else:
        import pandas as _pd
        import students as _students
        _batches = store.get("batches") or []
        if not _batches:
            st.warning("No batches in the store.")
        else:
            _b = st.selectbox("Batch", _batches,
                              index=len(_batches) - 1, key="stu_batch")
            _g = _store_grid(store["path"], _b, store["generated_at_iso"])
            if _g is None or _g.empty:
                st.warning(f"No roster grid stored for {_b}.")
            else:
                _summ = _students.summarise(list(_g.columns),
                                            _g.to_dict("records"))
                st.caption(
                    f"**{_summ['n_students']:,} students · {_summ['n_sessions']} "
                    "sessions.** A student is counted only against sessions they "
                    "were invited to — every whole-batch session plus their own "
                    "pod's. You cannot be absent from a session you were never "
                    "invited to, and dividing by every column is how a pod-heavy "
                    "batch comes to look like it collapsed."
                )
                c1, c2, c3, c4 = st.columns(4)
                c1.metric("Students", f"{_summ['n_students']:,}")
                c2.metric("Attended ≥1", f"{_summ['attended_any']:,}")
                c3.metric("Never attended", f"{_summ['never_attended']:,}")
                _bk = _summ["buckets"]
                c4.metric("Core (≥75%)", f"{_bk.get('core', 0):,}")

                st.subheader("How the batch splits")
                st.caption("An average hides whether everyone comes a third of "
                           "the time or a third of them come always — and those "
                           "need completely different responses.")
                _order = [("core", "Core ≥75%"), ("regular", "Regular 50-75%"),
                          ("occasional", "Occasional 25-50%"),
                          ("rare", "Rare <25%"), ("never", "Never attended")]
                st.dataframe(_pd.DataFrame(
                    [{"Group": lbl, "Students": _bk.get(k, 0),
                      "Share": (round(_bk.get(k, 0) / _summ["n_students"] * 100, 1)
                                if _summ["n_students"] else 0)}
                     for k, lbl in _order]), width='stretch', hide_index=True)

                st.subheader("Students")
                _q = st.text_input("Search email or phone", key="stu_q").strip().lower()
                _people = _summ["students"]
                if _q:
                    _people = [p for p in _people
                               if _q in str(p["email"]).lower()
                               or _q in str(p["phone"]).lower()]
                _mask = st.checkbox("Show full contact details", value=False,
                                    key="stu_mask")
                st.dataframe(_pd.DataFrame([{
                    "Email": p["email"] if _mask else mask_email(p["email"]),
                    "Phone": p["phone"] if _mask else mask_phone(p["phone"]),
                    "POD": p["pod"], "Active": p["active"],
                    "Attended": p["attended"], "Invited": p["invited"],
                    "Attendance %": p["pct"], "Unmarked": p["unmarked"],
                } for p in _people[:2000]]), width='stretch', hide_index=True,
                    height=460)
                if len(_people) > 2000:
                    st.caption(f"Showing the top 2,000 of {len(_people):,} by "
                               "attendance. Narrow with the search box — the "
                               "counts above always cover everyone.")
                if _summ["skipped_columns"]:
                    with st.expander("Columns not treated as sessions"):
                        st.write(", ".join(f"`{c}`" for c in _summ["skipped_columns"]))
                        st.caption("Matched on the date pattern, so a preference "
                                   "column can never be counted as a session and "
                                   "inflate everyone's denominator.")

# ============================ TAB 8 - BSIAI ==================================
with tab_bsiai:
    st.caption("The BSIAI programme. Separate roster, separate batch numbering - "
               "a session counts only once the L2 schedule lists its webinar.")

    _b = store.get("bsiai") if store_mode else None

    if not store_mode:
        st.info("The BSIAI tab reads the prebuilt store. It is empty in upload / "
                "legacy-live mode because BSIAI's roster is a different Sheet that "
                "the app does not fetch at runtime.")
    elif _b is None:
        st.warning(
            "**BSIAI is not configured yet.** Add `BSIAI_ROSTER_ID` (repo secret, "
            "or `bsiai_roster_id` under `[drive]` in secrets.toml) pointing at the "
            "BSIAI roster Sheet, and share that Sheet with the service account. "
            "The next refresh will fill this tab in."
        )
    elif not _b.get("DATA"):
        st.warning("BSIAI is configured but produced no batches in the last refresh.")
        for _w in _b.get("warnings", [])[:20]:
            st.caption("- " + str(_w))
    else:
        _note = (f"\U0001F7E2 Data as of {store['generated_at']} - {_b.get('source','')}"
                 .strip(" -"))
        dash_view.render(
            _b["DATA"], _b["summary"], _note, key="bsiai_batch",
            # BSIAI has no Close Type to slice by, so the same panel carries the
            # question its data CAN answer: how much of the course each learner
            # actually attends. A 55% average hides very different cohorts.
            show_closing=True,
            closing_title="Engagement",
            closing_sub=("bar = share of active learners · pill = that group's "
                         "own attendance rate"))
        _w = _b.get("warnings") or []
        if _w:
            with st.expander(f"Skipped / warnings ({len(_w)})"):
                for _line in _w:
                    st.text("- " + str(_line))

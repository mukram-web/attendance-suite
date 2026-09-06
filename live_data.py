"""
live_data.py — read the roster / L2 / attendee data straight from Google Drive.

The unified app uses this ONLY when live mode is configured (a service-account
key + Drive file IDs are present in Streamlit secrets). If anything here is
missing, the app silently falls back to manual file uploads — so the app always
runs, with or without Google set up.

Expected secrets (`.streamlit/secrets.toml`)::

    [gcp_service_account]
    type = "service_account"
    project_id = "..."
    private_key_id = "..."
    private_key = "-----BEGIN PRIVATE KEY-----\\n...\\n-----END PRIVATE KEY-----\\n"
    client_email = "robot@your-project.iam.gserviceaccount.com"
    client_id = "..."
    token_uri = "https://oauth2.googleapis.com/token"
    # ...the rest of the JSON key fields...

    [drive]
    roster_id       = "<google-sheet-file-id>"   # required
    l2_id           = "<google-sheet-file-id>"   # recommended (maps webinar -> batch)
    attendee_zip_id = "<zip-file-id>"            # the Zoom .zip you replace each week
    # OR, instead of a zip, a folder of attendee CSVs:
    attendee_folder_id = "<drive-folder-id>"

Only the `google-*` packages in requirements.txt are needed for this; they are
imported lazily so upload-only use never has to install them locally.
"""
from __future__ import annotations
import io
import os
import re
import socket
import threading
from concurrent.futures import ThreadPoolExecutor

# Attendee files on the Drive never change once uploaded, so their bytes are
# cached on disk by file id — restarts skip re-downloading hundreds of CSVs.
_CACHE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".cache", "attendee")
_SHEET_CACHE = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".cache", "sheets")

# Never let a single Drive request hang the whole app forever (the #1 cause of an
# app "stuck loading"). Any socket with no progress for this long raises instead.
socket.setdefaulttimeout(60)

# Drive/Sheets MIME types we care about
_SHEET_MIME = "application/vnd.google-apps.spreadsheet"
_FOLDER_MIME = "application/vnd.google-apps.folder"
_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]


def config_present() -> bool:
    """True only if a service account AND at least a roster id are configured."""
    try:
        import streamlit as st
        if "gcp_service_account" not in st.secrets:
            return False
        drive = st.secrets.get("drive", {})
        return bool(drive.get("roster_id"))
    except Exception:
        return False


# Optional credential override so this module also runs OUTSIDE Streamlit
# (pipeline.py on GitHub Actions). When unset, credentials come from st.secrets.
_CREDS_INFO: dict | None = None
_CREDS_SCOPES: list | None = None


def set_service_account(info: dict, scopes: list | None = None) -> None:
    """Use an explicit service-account JSON dict (and optionally wider scopes,
    e.g. drive read-write for the pipeline's uploads) instead of st.secrets."""
    global _CREDS_INFO, _CREDS_SCOPES
    _CREDS_INFO = dict(info)
    _CREDS_SCOPES = list(scopes) if scopes else None


def _drive_service():
    """Build a Drive v3 client — from the override if set, else st.secrets."""
    from google.oauth2.service_account import Credentials
    from googleapiclient.discovery import build

    if _CREDS_INFO is not None:
        info, scopes = dict(_CREDS_INFO), (_CREDS_SCOPES or _SCOPES)
    else:
        import streamlit as st
        info, scopes = dict(st.secrets["gcp_service_account"]), _SCOPES
    creds = Credentials.from_service_account_info(info, scopes=scopes)
    return build("drive", "v3", credentials=creds, cache_discovery=False)


def _download(request) -> bytes:
    """Run a Drive media request to completion and return the bytes."""
    from googleapiclient.http import MediaIoBaseDownload
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    return buf.getvalue()


def _file_meta(svc, file_id: str) -> dict:
    # supportsAllDrives=True so files living in a Shared Drive resolve too.
    return svc.files().get(fileId=file_id, fields="id, name, mimeType",
                           supportsAllDrives=True).execute()


def fetch_spreadsheet_xlsx(svc, file_id: str) -> bytes:
    """Return .xlsx bytes for a file id.

    If the id points at a native Google Sheet we EXPORT it to .xlsx; if it is an
    already-uploaded .xlsx we just download it. Either way the engines get bytes
    they can open with openpyxl.
    """
    meta = _file_meta(svc, file_id)
    if meta.get("mimeType") == _SHEET_MIME:
        req = svc.files().export_media(fileId=file_id, mimeType=_XLSX_MIME)
    else:
        req = svc.files().get_media(fileId=file_id, supportsAllDrives=True)
    return _download(req)


def fetch_file_bytes(svc, file_id: str) -> bytes:
    """Download any binary file (e.g. the attendee .zip) as raw bytes."""
    return _download(svc.files().get_media(fileId=file_id, supportsAllDrives=True))


def fetch_sheet_cached(svc, file_id: str) -> tuple[bytes, str]:
    """Like fetch_spreadsheet_xlsx, but disk-cached by the file's modifiedTime.

    Google's Sheet→xlsx export is NON-deterministic — exporting the same
    unchanged sheet twice returns different bytes — so downstream caches must
    key on the returned `stamp` (the Drive modifiedTime), never on the bytes.
    While the stamp is unchanged we also skip the export download entirely.
    Returns (xlsx_bytes, stamp)."""
    meta = svc.files().get(fileId=file_id, fields="modifiedTime, mimeType",
                           supportsAllDrives=True).execute()
    stamp = meta.get("modifiedTime", "")
    cpath = os.path.join(_SHEET_CACHE,
                         f"{file_id}_{re.sub(r'[^0-9A-Za-z]', '', stamp)}.xlsx")
    try:
        with open(cpath, "rb") as fh:
            return fh.read(), stamp
    except OSError:
        pass
    if meta.get("mimeType") == _SHEET_MIME:
        req = svc.files().export_media(fileId=file_id, mimeType=_XLSX_MIME)
    else:
        req = svc.files().get_media(fileId=file_id, supportsAllDrives=True)
    try:
        data = _download(req)
    except Exception as e:
        # Drive refuses to export a Google Sheet whose xlsx form exceeds 10 MB.
        # The roster is already ~8 MB and grows with every batch, so name the
        # cause instead of leaving a bare 403 in the logs.
        if "exportSizeLimitExceeded" in str(e) or "too large" in str(e).lower():
            raise RuntimeError(
                f"Google refused to export sheet {file_id} as .xlsx — it has grown "
                "past Drive's 10 MB export limit. The roster must be split (e.g. "
                "archive old batches into a separate sheet) before the pipeline "
                "can read it again."
            ) from e
        raise
    try:
        os.makedirs(_SHEET_CACHE, exist_ok=True)
        for old in os.listdir(_SHEET_CACHE):        # evict stale stamps of this file
            if old.startswith(file_id + "_"):
                try:
                    os.remove(os.path.join(_SHEET_CACHE, old))
                except OSError:
                    pass
        tmp = cpath + ".tmp"
        with open(tmp, "wb") as fh:
            fh.write(data)
        os.replace(tmp, cpath)
    except OSError:
        pass
    return data, stamp


# Per-thread Drive client — googleapiclient's http transport is not safe to share
# across threads, so each worker builds its own (cheap; discovery is bundled).
_thread_local = threading.local()


def _thread_drive():
    svc = getattr(_thread_local, "svc", None)
    if svc is None:
        svc = _drive_service()
        _thread_local.svc = svc
    return svc


def _folder_ids(folder_id) -> list[str]:
    """Split a comma-separated attendee-folder config value into individual ids.
    Sessions can live on more than one Shared Drive (the original one plus the
    'Zoom extracts' drive new sessions are written to since Aug 2026)."""
    return [p.strip() for p in str(folder_id or "").split(",") if p.strip()]


def _list_children(svc, folder_id: str) -> list[dict]:
    """One folder's immediate children (id, name, mimeType), Shared-Drive aware."""
    items, page_token = [], None
    while True:
        resp = svc.files().list(
            q=f"'{folder_id}' in parents and trashed=false",
            # `size` matters: duplicate session folders exist on this drive and
            # callers break the tie on it. Drive only returns requested fields,
            # so leaving it out silently makes every file look 0 bytes.
            fields="nextPageToken, files(id, name, mimeType, size)",
            pageSize=1000, pageToken=page_token,
            supportsAllDrives=True, includeItemsFromAllDrives=True,
        ).execute()
        items += resp.get("files", [])
        page_token = resp.get("nextPageToken")
        if not page_token:
            break
    return items


def list_attendee_names(folder_id: str | None = None) -> list:
    """All attendee filenames across the Shared Drive — names only, no downloads.

    One Drive query (`name contains 'attendee_'`) over the whole Shared Drive, so
    it's fast even with 500+ session folders. The filenames carry the Webinar ID
    and date, which is all the dashboard's topic join needs. Falls back to a
    recursive name-only walk for a plain (non-Shared-Drive) folder.
    """
    fid = folder_id
    if fid is None:
        import streamlit as st
        fid = st.secrets["drive"].get("attendee_folder_id")
    if not fid:
        return []
    svc = _drive_service()
    names: list = []
    for one in _folder_ids(fid):
        if one.startswith("0A"):                        # Shared Drive root → whole-drive search
            token = None
            while True:
                resp = svc.files().list(
                    q="name contains 'attendee_' and trashed = false",
                    corpora="drive", driveId=one,
                    includeItemsFromAllDrives=True, supportsAllDrives=True,
                    fields="nextPageToken, files(name)", pageSize=1000, pageToken=token,
                ).execute()
                names += [f["name"] for f in resp.get("files", [])]
                token = resp.get("nextPageToken")
                if not token:
                    break
        else:                                           # plain folder → recurse names only
            def walk(f):
                for c in _list_children(svc, f):
                    if c["mimeType"] == _FOLDER_MIME:
                        walk(c["id"])
                    elif "attendee" in c["name"].lower():
                        names.append(c["name"])
            walk(one)
    return names


def find_in_folder(svc, folder_id: str, name: str) -> dict | None:
    """First non-trashed file with this exact name inside the folder (Shared-Drive
    aware). Returns its metadata dict or None."""
    safe = name.replace("\\", "\\\\").replace("'", "\\'")
    resp = svc.files().list(
        q=f"'{folder_id}' in parents and name = '{safe}' and trashed = false",
        fields="files(id, name, mimeType, modifiedTime, size)", pageSize=10,
        supportsAllDrives=True, includeItemsFromAllDrives=True,
    ).execute()
    files = resp.get("files", [])
    return files[0] if files else None


def upload_to_folder(svc, folder_id: str, name: str, data: bytes,
                     mime: str = "application/octet-stream") -> str:
    """Create-or-replace `name` inside the folder with these bytes; returns the
    file id. Needs the service account to have Editor access on the folder and
    a read-write Drive scope (see set_service_account)."""
    from googleapiclient.http import MediaIoBaseUpload
    media = MediaIoBaseUpload(io.BytesIO(data), mimetype=mime, resumable=True)
    existing = find_in_folder(svc, folder_id, name)
    if existing:
        return svc.files().update(fileId=existing["id"], media_body=media,
                                  supportsAllDrives=True).execute().get("id", existing["id"])
    meta = {"name": name, "parents": [folder_id]}
    return svc.files().create(body=meta, media_body=media, fields="id",
                              supportsAllDrives=True).execute()["id"]


def _download_any(svc, file_id: str) -> bytes:
    """get_media, falling back to a text/csv export for Google-native files —
    attendee reports occasionally get converted to a Google Sheet on upload,
    which the binary endpoint refuses with 403 fileNotDownloadable."""
    try:
        return _download(svc.files().get_media(fileId=file_id,
                                               supportsAllDrives=True))
    except Exception as e:
        if "fileNotDownloadable" not in str(e):
            raise
        return _download(svc.files().export_media(fileId=file_id,
                                                  mimeType="text/csv"))


def download_cached(svc, file_id: str) -> bytes:
    """Download a Drive file, reusing the per-file-id byte cache. Attendee CSVs
    never change once uploaded, so a second reader (e.g. the day-1 analysis)
    costs nothing after the marker has already pulled them."""
    cpath = os.path.join(_CACHE_DIR, file_id)
    try:
        with open(cpath, "rb") as fh:
            return fh.read()
    except OSError:
        pass
    data = _download_any(svc, file_id)
    try:
        os.makedirs(_CACHE_DIR, exist_ok=True)
        tmp = cpath + ".tmp"
        with open(tmp, "wb") as fh:
            fh.write(data)
        os.replace(tmp, cpath)
    except OSError:
        pass
    return data


def fetch_store(store_folder_id: str, name: str = "attendance.duckdb"):
    """Download the prebuilt dashboard store from the Drive folder the pipeline
    writes to. Returns (bytes, modifiedTime) or (None, None) if absent."""
    svc = _drive_service()
    meta = find_in_folder(svc, store_folder_id, name)
    if not meta:
        return None, None
    data = _download(svc.files().get_media(fileId=meta["id"], supportsAllDrives=True))
    return data, meta.get("modifiedTime", "")


def list_store_snapshots(store_folder_id: str) -> list[dict]:
    """Past weeks' stores from archive/, newest first.

    Each is {"name", "id", "date", "size"}. Returns [] when the folder or the
    archive does not exist yet, so a caller can offer history when there is some
    and stay quiet when there is not.
    """
    import archive
    svc = _drive_service()
    try:
        folder = archive.ensure_folder(svc, store_folder_id)
    except Exception:
        return []
    out = []
    for f in _list_children(svc, folder):
        name = f.get("name", "")
        m = re.match(r"attendance_store_(\d{4}-\d{2}-\d{2})\.duckdb$", name)
        if m:
            out.append({"name": name, "id": f["id"], "date": m.group(1),
                        "size": int(f.get("size") or 0)})
    return sorted(out, key=lambda d: d["date"], reverse=True)


def find_archived(store_folder_id: str, name: str) -> dict | None:
    """One named file inside archive/, or None. Used to fetch a PAST week's
    marked workbook, which lives only under its dated name."""
    import archive
    svc = _drive_service()
    try:
        return find_in_folder(svc, archive.ensure_folder(svc, store_folder_id), name)
    except Exception:
        return None


def fetch_store_snapshot(file_id: str) -> bytes:
    """One archived store by file id. Snapshots are immutable, so the byte cache
    is always valid — reopening last month's week costs nothing after the first."""
    return download_cached(_drive_service(), file_id)


def _existing_sessions(roster_bytes: bytes) -> dict:
    """{batch_key -> set(mm_dd)} of session columns already present in the roster,
    so we can skip re-fetching sessions that are already marked.

    Uses openpyxl read-only mode and reads only the top rows of each sheet — far
    lighter on memory than a full load (matters on Streamlit's small instances).
    """
    import attendance_core as ac
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(roster_bytes), read_only=True, data_only=True)

    def _has(row, needle):
        return any(isinstance(v, str) and needle in v.strip().lower() for v in row)

    out: dict = {}
    for name in wb.sheetnames:
        k = ac._sheet_key(name)
        if not k:
            continue
        ws = wb[name]
        top = list(ws.iter_rows(min_row=1, max_row=3, max_col=200, values_only=True))
        row1 = top[0] if top else ()
        row2 = top[1] if len(top) > 1 else ()
        hr = 2 if (_has(row2, "registered number") or _has(row2, "payment")) else 1
        dates = set()
        width = max(len(row1), len(row2))
        for c in range(10, width):              # 0-based: session columns start at K (index 10)
            v1 = row1[c] if c < len(row1) else None
            v2 = row2[c] if (hr == 2 and c < len(row2)) else None
            d = ac._mmdd(v1) or (ac._mmdd(v2) if hr == 2 else None)
            if d:
                dates.add(d)
        out[k] = dates
    wb.close()
    return out


def fetch_new_attendees(svc, folder_id: str, roster_bytes: bytes,
                        mark_all: bool = False, max_workers: int = 8):
    """Download attendee files only for sessions NOT already marked in the roster.

    The Shared Drive holds the full history (hundreds of dated session folders),
    but the roster already carries every past weekend's marks. So we list the
    top-level dated folders, keep only those whose (batch, date) is missing from
    the roster, and download just those — in parallel. `mark_all=True` ignores the
    skip logic and pulls everything (a full rebuild). Returns (attendee_files, info).
    """
    import io as _io
    import zipfile
    import attendance_core as ac

    existing = _existing_sessions(roster_bytes)
    sheet_keys = set(existing)

    def keep(name: str) -> bool:
        """Attendee reports only. A .zip qualifies solely when its name says it
        holds attendee data — otherwise a session-recording zip dropped into a
        folder would be downloaded whole (potentially gigabytes) for nothing."""
        n = name.lower()
        return n.startswith("attendee") or (n.endswith(".zip") and "attendee" in n)

    # 1) pick the top-level folders that need marking — across every configured
    #    drive (comma-separated ids: the original drive plus any newer one)
    to_fetch, skipped_nosheet, skipped_done = [], 0, 0
    top: list[dict] = []
    for fid in _folder_ids(folder_id):
        top += _list_children(svc, fid)
    for f in top:
        if f["mimeType"] != _FOLDER_MIME:
            continue
        name = f["name"]
        mm = ac._mmdd(name)
        covered = [k for k in ac._folder_batches(name) if k in sheet_keys]
        if not covered:
            skipped_nosheet += 1
            continue
        if mark_all or mm is None or any(mm not in existing[k] for k in covered):
            to_fetch.append((name, f["id"]))
        else:
            skipped_done += 1

    # 2) collect the attendee files inside those folders — in parallel: one API
    #    call per folder, and at 200+ unmarked folders doing them sequentially
    #    dominates cold-start time
    entries = []  # (path, file_id, is_zip)
    listing_errors = []
    if to_fetch:
        def _kids(item):
            name, fid = item
            try:
                return name, _list_children(_thread_drive(), fid), None
            except Exception as e:
                return name, [], str(e)
        with ThreadPoolExecutor(max_workers=min(max_workers, len(to_fetch))) as ex:
            for name, kids, kerr in ex.map(_kids, to_fetch):
                if kerr:
                    # A folder we couldn't list is a session we'd silently drop —
                    # surface it so the caller can refuse to publish partial data.
                    listing_errors.append(f"{name}: {kerr}")
                    continue
                for f in kids:
                    if f["mimeType"] != _FOLDER_MIME and keep(f["name"]):
                        entries.append((f"{name}/{f['name']}", f["id"],
                                        f["name"].lower().endswith(".zip")))

    # 3) download (parallel, but resilient — one slow/failed file can't hang or
    #    sink the whole batch; the 60s socket timeout caps any single request).
    #    Bytes are disk-cached by file id, so only never-seen files hit the network.
    out: list[tuple[str, bytes]] = []
    failed = 0
    bad_zips: list[str] = []
    download_errors: list[str] = []
    if entries:
        try:
            os.makedirs(_CACHE_DIR, exist_ok=True)
        except OSError:
            pass

        def _dl(e):
            path, fid, is_zip = e
            cpath = os.path.join(_CACHE_DIR, fid)
            try:
                with open(cpath, "rb") as fh:
                    return path, fid, is_zip, fh.read(), None
            except OSError:
                pass
            try:
                data = _download_any(_thread_drive(), fid)
                try:
                    tmp = cpath + ".tmp"
                    with open(tmp, "wb") as fh:
                        fh.write(data)
                    os.replace(tmp, cpath)
                except OSError:
                    pass
                return path, fid, is_zip, data, None
            except Exception as e:
                return path, fid, is_zip, None, str(e)
        with ThreadPoolExecutor(max_workers=min(max_workers, len(entries))) as ex:
            for path, fid, is_zip, data, err in ex.map(_dl, entries):
                if data is None:
                    failed += 1
                    download_errors.append(f"{path}: {err}")
                    continue
                if is_zip:
                    # A corrupt / password-protected / non-attendee .zip must not
                    # sink the whole run — and its bytes must not stay cached, or
                    # every later run would fail identically without a network call.
                    prefix = (path.rsplit("/", 1)[0] + "/") if "/" in path else ""
                    try:
                        with zipfile.ZipFile(_io.BytesIO(data)) as z:
                            expanded = [(f"{prefix}{inner}", z.read(inner))
                                        for inner in z.namelist()
                                        if not inner.endswith("/")]
                    except Exception as e:
                        failed += 1
                        bad_zips.append(f"{path}: {e}")
                        try:
                            os.remove(os.path.join(_CACHE_DIR, fid))
                        except OSError:
                            pass
                        continue
                    out.extend(expanded)
                else:
                    out.append((path, data))

    info = dict(new_folders=len(to_fetch), files=len(out), failed=failed,
                skipped_already_marked=skipped_done, skipped_no_sheet=skipped_nosheet,
                listing_errors=listing_errors, bad_zips=bad_zips,
                download_errors=download_errors)
    return out, info


def fetch_track_attendees(svc, folder_id, track: str, max_workers: int = 8):
    """Every attendee report belonging to one programme track (e.g. "BSIAI").

    `fetch_new_attendees` selects folders by which roster sheet already carries
    which dates — that only works for a programme whose workbook holds session
    columns. BSIAI has none: its attendance is recomputed from the reports each
    run, so this simply takes every folder whose name resolves to the requested
    track, across every configured drive.

    Cheap despite that: downloads are disk-cached by Drive file id, so a re-run
    touches the network only for genuinely new files.

    Returns (attendee_files, info) with the same info keys the caller's
    partial-data gate already understands.
    """
    import attendance_core as ac

    def keep(name: str) -> bool:
        return name.lower().startswith("attendee")

    folders = []
    for fid in _folder_ids(folder_id):
        for f in _list_children(svc, fid):
            if f["mimeType"] != _FOLDER_MIME:
                continue
            if any(t == track for t, _ in ac._folder_batches(f["name"])):
                folders.append((f["name"], f["id"]))

    entries, listing_errors = [], []
    if folders:
        def _kids(item):
            name, fid = item
            try:
                return name, _list_children(_thread_drive(), fid), None
            except Exception as e:
                return name, [], str(e)
        with ThreadPoolExecutor(max_workers=min(max_workers, len(folders))) as ex:
            for name, kids, kerr in ex.map(_kids, folders):
                if kerr:
                    listing_errors.append(f"{name}: {kerr}")
                    continue
                for f in kids:
                    if f["mimeType"] != _FOLDER_MIME and keep(f["name"]):
                        entries.append((f"{name}/{f['name']}", f["id"]))

    out, failed, download_errors = [], 0, []
    if entries:
        try:
            os.makedirs(_CACHE_DIR, exist_ok=True)
        except OSError:
            pass

        def _dl(e):
            path, fid = e
            cpath = os.path.join(_CACHE_DIR, fid)
            try:
                with open(cpath, "rb") as fh:
                    return path, fh.read(), None
            except OSError:
                pass
            try:
                data = _download_any(_thread_drive(), fid)
                try:
                    tmp = cpath + ".tmp"
                    with open(tmp, "wb") as fh:
                        fh.write(data)
                    os.replace(tmp, cpath)
                except OSError:
                    pass
                return path, data, None
            except Exception as e:
                return path, None, str(e)

        with ThreadPoolExecutor(max_workers=min(max_workers, len(entries))) as ex:
            for path, data, err in ex.map(_dl, entries):
                if data is None:
                    failed += 1
                    download_errors.append(f"{path}: {err}")
                    continue
                out.append((path, data))

    info = dict(new_folders=len(folders), files=len(out), failed=failed,
                skipped_already_marked=0, skipped_no_sheet=0,
                listing_errors=listing_errors, bad_zips=[],
                download_errors=download_errors)
    return out, info


def scan_all_attendees(svc, folder_id, consume, max_workers: int = 8):
    """Download EVERY attendee report and hand each to `consume(name, bytes)`.

    Unlike `fetch_new_attendees`, this covers the whole history — the duration
    and peak columns need the report for every session, not just the unmarked
    ones. Measured 2026-09-06: 1,284 files, 246 MB, and the pipeline already
    downloads at ~7.9 files/sec, so about 2.7 minutes on top of a 7.5-minute
    run. Well inside the workflow's 30-minute timeout.

    **Nothing is accumulated.** `consume` is called with each file's bytes and
    the bytes are dropped immediately, because holding 246 MB is exactly how the
    marking used to blow up on a small instance. Callers keep the small
    aggregate, never the corpus.

    Failures are counted, never raised: a missing report should cost that
    session its duration, not the whole refresh.
    """
    files, seen = [], set()
    for did in _folder_ids(folder_id):
        token = None
        while True:
            kw = dict(q="name contains 'attendee_' and trashed = false",
                      fields="nextPageToken, files(id, name)", pageSize=1000,
                      pageToken=token, includeItemsFromAllDrives=True,
                      supportsAllDrives=True)
            if str(did).startswith("0A"):
                kw.update(corpora="drive", driveId=did)
            resp = svc.files().list(**kw).execute()
            for f in resp.get("files", []):
                if f["id"] not in seen:
                    seen.add(f["id"])
                    files.append(f)
            token = resp.get("nextPageToken")
            if not token:
                break

    failed = 0
    if not files:
        return {"files": 0, "failed": 0}
    try:
        os.makedirs(_CACHE_DIR, exist_ok=True)
    except OSError:
        pass

    def _dl(f):
        cpath = os.path.join(_CACHE_DIR, f["id"])
        try:
            with open(cpath, "rb") as fh:
                return f["name"], fh.read(), None
        except OSError:
            pass
        try:
            data = _download_any(_thread_drive(), f["id"])
            try:
                tmp = cpath + ".tmp"
                with open(tmp, "wb") as fh:
                    fh.write(data)
                os.replace(tmp, cpath)
            except OSError:
                pass
            return f["name"], data, None
        except Exception as e:
            return f["name"], None, str(e)

    done = 0
    with ThreadPoolExecutor(max_workers=min(max_workers, len(files))) as ex:
        for name, data, err in ex.map(_dl, files):
            if data is None:
                failed += 1
                continue
            try:
                consume(name, data)
            except Exception:
                failed += 1
            finally:
                del data          # keep peak memory at one file, not 246 MB
            done += 1
    return {"files": done, "failed": failed}


def fetch_polls(svc, folder_id, max_workers: int = 8):
    """Every `poll_<webinarid>_<date>.csv` across the configured drives.

    One name query per drive rather than a walk - there are 1,100+ of them - and
    the same disk cache the attendee reports use, so a re-run costs nothing.
    Failures are counted, never raised: a missing poll should cost that session
    its rating, not the whole refresh.
    """
    files, seen = [], set()
    for did in _folder_ids(folder_id):
        token = None
        while True:
            # both conventions - see polls.name_key
            kw = dict(q=("(name contains 'poll_' or name contains 'Poll Report') "
                         "and trashed = false"),
                      fields="nextPageToken, files(id, name)", pageSize=1000,
                      pageToken=token, includeItemsFromAllDrives=True,
                      supportsAllDrives=True)
            if str(did).startswith("0A"):
                kw.update(corpora="drive", driveId=did)
            resp = svc.files().list(**kw).execute()
            for f in resp.get("files", []):
                if f["id"] not in seen:
                    seen.add(f["id"])
                    files.append(f)
            token = resp.get("nextPageToken")
            if not token:
                break

    out, failed = [], 0
    if files:
        try:
            os.makedirs(_CACHE_DIR, exist_ok=True)
        except OSError:
            pass

        def _dl(f):
            cpath = os.path.join(_CACHE_DIR, f["id"])
            try:
                with open(cpath, "rb") as fh:
                    return f["name"], fh.read(), None
            except OSError:
                pass
            try:
                data = _download_any(_thread_drive(), f["id"])
                try:
                    tmp = cpath + ".tmp"
                    with open(tmp, "wb") as fh:
                        fh.write(data)
                    os.replace(tmp, cpath)
                except OSError:
                    pass
                return f["name"], data, None
            except Exception as e:
                return f["name"], None, str(e)

        with ThreadPoolExecutor(max_workers=min(max_workers, len(files))) as ex:
            for name, data, err in ex.map(_dl, files):
                if data is None:
                    failed += 1
                    continue
                out.append((name, data))
    return out, {"found": len(files), "files": len(out), "failed": failed}


def load_live():
    """Pull everything the marker needs from Drive.

    Returns a dict::
        {
          "roster_bytes": bytes,
          "l2_bytes": bytes | None,
          "attendee_files": [(name, bytes), ...] | None,   # None if no zip/folder set
          "source": "<human description>",
        }
    Raises on any Drive/auth error so the caller can show a clear message and
    offer the upload fallback.
    """
    import streamlit as st
    drive = st.secrets["drive"]
    svc = _drive_service()

    roster_bytes, roster_stamp = fetch_sheet_cached(svc, drive["roster_id"])

    l2_bytes, l2_stamp = None, ""
    if drive.get("l2_id"):
        l2_bytes, l2_stamp = fetch_sheet_cached(svc, drive["l2_id"])

    attendee_files = None
    info = {}
    source_bits = ["roster"]
    if drive.get("attendee_zip_id"):
        import zipfile
        zb = fetch_file_bytes(svc, drive["attendee_zip_id"])
        with zipfile.ZipFile(io.BytesIO(zb)) as z:
            attendee_files = [(n, z.read(n)) for n in z.namelist() if not n.endswith("/")]
        source_bits.append(f"{len(attendee_files)} files from attendee .zip")
    elif drive.get("attendee_folder_id"):
        mark_all = bool(drive.get("mark_all", False))
        attendee_files, info = fetch_new_attendees(
            svc, drive["attendee_folder_id"], roster_bytes, mark_all=mark_all)
        if mark_all:
            source_bits.append(f"{info['files']} files (full rebuild)")
        else:
            source_bits.append(f"{info['files']} file(s) from {info['new_folders']} new session(s)")

    return {
        "roster_bytes": roster_bytes,
        "l2_bytes": l2_bytes,
        "attendee_files": attendee_files,
        "info": info,
        "source": "Google Drive — " + ", ".join(source_bits),
        # modifiedTime stamps: the stable identity for downstream caches
        # (export bytes vary run-to-run even when the sheet is unchanged)
        "roster_stamp": roster_stamp,
        "l2_stamp": l2_stamp,
    }

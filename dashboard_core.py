"""
Be10X Attendance Dashboard — compute engine.

Reads a Master Batch Roster (.xlsx) and returns a tidy attendance table.
The roster already holds Present/Absent per session column and a Payment
column we use to decide who is an "active" student.

Public API:
    compute(roster_bytes: bytes) -> pandas.DataFrame
    batch_key(name: str) -> int          # for natural batch sorting

Returned DataFrame columns:
    Batch, SessionIdx, SessionLabel, Topic,
    PresentActive, Active, PresentAll, Total,
    PctActive, PctAll, HasData
"""
from __future__ import annotations
import io
import re
import openpyxl
import pandas as pd

# Payment values that mean the student is NOT active (refunded / unidentified / blank)
_REFUND_HINTS = ("refund", "undifined", "unidentif", "undefined")


def _is_active(payment) -> bool:
    if payment is None:
        return False
    p = str(payment).strip().lower()
    if p == "":
        return False
    return not any(h in p for h in _REFUND_HINTS)


def batch_key(name: str) -> int:
    """Natural sort key: 'B17' -> 17, 'AI CAP B7' -> 7."""
    m = re.sub(r"\D", "", str(name))
    return int(m) if m else 0


def _find_header_row(ws):
    """Return (row_number, lowercased_values) for the row that holds field names."""
    for r in range(1, 4):
        row = next(ws.iter_rows(min_row=r, max_row=r, max_col=20))
        vals = [str(c.value).strip().lower() if c.value is not None else "" for c in row]
        if "registered number" in vals or "payment" in vals:
            return r, vals
    return 1, []


def _col_idx(vals, name):
    for i, v in enumerate(vals):
        if v == name:
            return i
    return None


def _looks_like_roster(sheet_name: str) -> bool:
    s = sheet_name.lower()
    if "att" in s:          # skip Zoom "...Att" helper sheets
        return False
    return True


def compute(roster_bytes: bytes) -> pd.DataFrame:
    wb = openpyxl.load_workbook(io.BytesIO(roster_bytes), read_only=True, data_only=True)
    records = []

    for sh in wb.sheetnames:
        if not _looks_like_roster(sh):
            continue
        ws = wb[sh]

        top = list(ws.iter_rows(min_row=1, max_row=2, max_col=80, values_only=True))
        row1 = top[0] if len(top) > 0 else ()
        row2 = top[1] if len(top) > 1 else ()

        H, hvals = _find_header_row(ws)
        pay_i = _col_idx(hvals, "payment")
        close_i = _col_idx(hvals, "closing type")
        if close_i is None:
            close_i = (pay_i + 1) if pay_i is not None else 9

        # Session columns live to the right of "Closing Type".
        # The date label for each session is always in row 1 at that column.
        sess_cols = []
        for c in range(close_i + 1, len(row1)):
            label = row1[c] if c < len(row1) else None
            if label is None or str(label).strip() == "":
                continue
            topic = ""
            if H == 2 and c < len(row2) and row2[c] is not None:
                topic = str(row2[c]).strip()
            sess_cols.append((c, str(label).strip(), topic))

        if not sess_cols:
            continue

        last_col = max([c for c, _, _ in sess_cols] + [pay_i or 0, close_i]) + 1
        data = list(ws.iter_rows(min_row=H + 1, max_col=last_col, values_only=True))

        active = 0
        total = 0
        p_active = [0] * len(sess_cols)
        p_all = [0] * len(sess_cols)

        for row in data:
            if all((x is None or str(x).strip() == "") for x in row[:8]):
                continue
            total += 1
            act = _is_active(row[pay_i]) if (pay_i is not None and pay_i < len(row)) else False
            if act:
                active += 1
            for k, (c, _lbl, _t) in enumerate(sess_cols):
                v = row[c] if c < len(row) else None
                if v is not None and str(v).strip().lower() == "present":
                    p_all[k] += 1
                    if act:
                        p_active[k] += 1

        bn = sh.replace("AI CAP", "").replace("AICAP", "").strip()
        for k, (c, lbl, topic) in enumerate(sess_cols):
            has_data = p_all[k] > 0
            records.append(
                dict(
                    Batch=bn,
                    SessionIdx=k + 1,
                    SessionLabel=lbl,
                    Topic=topic,
                    PresentActive=p_active[k],
                    Active=active,
                    PresentAll=p_all[k],
                    Total=total,
                    PctActive=round(p_active[k] / active * 100, 1) if active else 0.0,
                    PctAll=round(p_all[k] / total * 100, 1) if total else 0.0,
                    HasData=has_data,
                )
            )

    df = pd.DataFrame.from_records(records)
    if not df.empty:
        df = df.sort_values(by=["Batch", "SessionIdx"], key=lambda s: s.map(batch_key) if s.name == "Batch" else s)
        df = df.reset_index(drop=True)
    return df


def _clean_batch_name(sheet_name: str) -> str:
    return sheet_name.replace("AI CAP", "").replace("AICAP", "").strip()


def batch_sheet_map(roster_bytes: bytes) -> dict:
    """Map cleaned batch name (e.g. 'B17') -> actual sheet name (e.g. 'AI CAP B17')."""
    wb = openpyxl.load_workbook(io.BytesIO(roster_bytes), read_only=True)
    out = {}
    for sh in wb.sheetnames:
        if _looks_like_roster(sh):
            out.setdefault(_clean_batch_name(sh), sh)
    wb.close()
    return out


def roster_grid(roster_bytes: bytes, sheet_name: str) -> pd.DataFrame:
    """Per-student attendance grid for ONE batch sheet — the spreadsheet view.

    Columns: Email, Phone, Active, Present (count), then one column per session
    labelled by its date, holding 'Present' / 'Absent' / '' exactly as marked.
    Contact columns are raw here; the UI masks them for privacy.
    """
    wb = openpyxl.load_workbook(io.BytesIO(roster_bytes), read_only=True, data_only=True)
    ws = wb[sheet_name]

    top = list(ws.iter_rows(min_row=1, max_row=2, max_col=80, values_only=True))
    row1 = top[0] if len(top) > 0 else ()

    H, hvals = _find_header_row(ws)
    pay_i = _col_idx(hvals, "payment")
    close_i = _col_idx(hvals, "closing type")
    if close_i is None:
        close_i = (pay_i + 1) if pay_i is not None else 9
    mail_i = _col_idx(hvals, "registered mail")
    num_i = _col_idx(hvals, "registered number")

    sess_cols = []
    for c in range(close_i + 1, len(row1)):
        label = row1[c] if c < len(row1) else None
        if label is None or str(label).strip() == "":
            continue
        sess_cols.append((c, str(label).strip()))

    last_col = max([c for c, _ in sess_cols] + [pay_i or 0, close_i, mail_i or 0, num_i or 0]) + 1
    data = list(ws.iter_rows(min_row=H + 1, max_col=last_col, values_only=True))
    wb.close()

    records = []
    for row in data:
        if all((x is None or str(x).strip() == "") for x in row[:8]):
            continue
        email = row[mail_i] if (mail_i is not None and mail_i < len(row)) else None
        phone = row[num_i] if (num_i is not None and num_i < len(row)) else None
        active = _is_active(row[pay_i]) if (pay_i is not None and pay_i < len(row)) else False
        rec = {
            "Email": "" if email is None else str(email).strip(),
            "Phone": "" if phone is None else str(phone).strip().replace(".0", ""),
            "Active": active,
        }
        present = 0
        for c, label in sess_cols:
            v = row[c] if c < len(row) else None
            mark = "" if v is None else str(v).strip().title()
            rec[label] = mark
            if mark.lower() == "present":
                present += 1
        rec["Present"] = present
        records.append(rec)

    cols = ["Email", "Phone", "Active", "Present"] + [lbl for _, lbl in sess_cols]
    return pd.DataFrame.from_records(records, columns=cols)

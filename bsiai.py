"""
bsiai.py — the BSIAI programme's attendance, computed independently of AI CAP.

Why a separate module rather than another branch inside `data.py`:

* **Different roster shape.** BSIAI's roster is its own Sheet with columns
  `CountryCode / Phone / CC+Phone / Email / Batch / … / Refund`. There is no
  `Payment` and no `Close Type`, so neither the active/inactive vocabulary nor
  the closing-type panel that `data.py` is built around applies here.
* **Different batch naming.** The roster says `BSIAI B1GA`, `BSI B2W1`; Drive
  says `BSI B1`, `bsi b2`, `BSIAI B1`; L2 says `BSIAI B1`. One normaliser
  (`batch_num`) reconciles all three.
* **No session columns anywhere.** AI CAP carries Present/Absent columns in the
  workbook and the marker appends new ones. BSIAI has none, so attendance is
  computed straight from the attendee reports against the roster — there is no
  workbook to mark and nothing is ever written back.

Rules (agreed with the owner, 2026-08-28):

1. **A session counts only if L2 knows its webinar id.** The AI CAP path falls
   back to the folder name when a webinar is missing from L2; BSIAI deliberately
   does not. No L2 row → the session is skipped and listed as a warning.
2. **Strength excludes refunds** (`Refund` cell truthy). There is no other
   status column, so "Active" here means "enrolled and not refunded".
3. **Batches are discovered from the data.** A batch appears once it has both
   roster rows and at least one L2-known session; nothing is hardcoded.

The output is deliberately the same shape `data.build` returns, so `dash_view`
renders it with no special-casing (`closing` is empty and the caller passes
`show_closing=False`).
"""
from __future__ import annotations

import io
import re
from collections import defaultdict

from attendance_core import _digits, _mmdd, parse_attendees
from data import date_label

# Roster batch labels carry a group suffix the shared parser cannot read:
# 'BSIAI B1GA' (group A), 'BSI B2W1' (week 1). Both belong to the parent batch.
_BATCH_RE = re.compile(r"b\s*(\d{1,2})\s*(?:g[a-z]|w\d)?\s*$", re.I)
_TRUE = {"true", "yes", "y", "1", "1.0", "refund", "refunded"}

# Tabs in the BSIAI sheet that are deliberately NOT roster data. Confirmed with
# the owner on 2026-08-29, so they are skipped silently rather than reported as
# warnings every week:
#   failed BSI B1        - a Q&A group split (Group A / Group B email lists)
#   Sheet8               - contacts with no Batch column, unassignable to a batch
#   NEXT BATCH 15K MMAI  - MM-AI B1, a different programme
# Matched case-insensitively on the trimmed name. Anything NOT on this list still
# warns - that is what stops a real roster tab from going missing in silence, so
# add to it only for tabs someone has actually looked at.
_IGNORED_TABS = {"failed bsi b1", "sheet8", "next batch 15k mmai"}


def batch_num(label) -> int | None:
    """Any BSIAI batch label -> its number. None if it is not a BSIAI batch.

    Handles every spelling seen in the three sources:
    'BSIAI B1', 'BSIAI B1GA', 'BSI AI B3', 'BSI B1', 'bsi b2', 'BSI B2W1'.
    'MM-AI B1' is a different programme and returns None.
    """
    s = re.sub(r"\s+", " ", str(label or "")).strip()
    if not s:
        return None
    if not re.search("(?<![a-z])bsi", s, re.I):
        return None
    m = _BATCH_RE.search(s)
    return int(m.group(1)) if m else None


def is_refunded(cell) -> bool:
    return str(cell or "").strip().lower() in _TRUE


def _norm_email(v) -> str:
    return re.sub(r"\s", "", str(v or "")).lower()


def _norm_phone(v) -> str:
    """Digits only, with openpyxl's float tail removed FIRST.

    Stripping non-digits before the '.0' turns 915550000101.0 into
    9155500001010 — a silently shifted number that never matches again.
    """
    return _digits(re.sub(r"\.0$", "", str(v or "").strip()))


# ─────────────────────────────── roster ──────────────────────────────────────
def read_roster(xlsx_bytes: bytes) -> tuple[dict, list]:
    """BSIAI roster workbook -> ({batch_num: [student, …]}, warnings).

    Every worksheet is scanned; one is treated as roster data if its header row
    names both an email and a batch column. Tabs that do not (pivots, scratch,
    the MM-AI portal export) are skipped with a warning rather than crashing —
    people add tabs, and a stray one must never take the build down.

    Tabs named in `_IGNORED_TABS` are skipped silently; any other unusable tab
    still produces a warning, so a real roster tab can never vanish unnoticed.

    student = {"email": str, "phone": str, "refund": bool, "label": str}
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    out: dict[int, list] = defaultdict(list)
    warnings: list[str] = []

    for ws in wb.worksheets:
        if ws.title.strip().lower() in _IGNORED_TABS:
            continue                      # known non-roster tab - not a warning
        rows = ws.iter_rows(max_col=40, values_only=True)
        header_i = cols = None
        for i, row in enumerate(rows):
            if i > 8:
                break
            low = [str(c or "").strip().lower() for c in row]
            if any("email" in c for c in low) and any(c == "batch" or c.startswith("batch")
                                                      for c in low):
                header_i = i
                cols = low
                break
        if cols is None:
            warnings.append(f'Tab "{ws.title}": no Email/Batch header - skipped')
            continue

        def find(*needles, exclude=()):
            for j, h in enumerate(cols):
                if all(n in h for n in needles) and not any(x in h for x in exclude):
                    return j
            return None

        # 'lms email' must not win over the real 'Email' column.
        e_i = next((j for j, h in enumerate(cols) if h == "email"), None)
        if e_i is None:
            e_i = find("email", exclude=("lms",))
        p_i = find("cc+phone") or find("phone", exclude=("lms",))
        b_i = next((j for j, h in enumerate(cols) if h.startswith("batch")), None)
        r_i = find("refund")

        n_before = sum(len(v) for v in out.values())
        for row in ws.iter_rows(min_row=header_i + 2, max_col=40, values_only=True):
            if row is None:
                continue
            email = _norm_email(row[e_i]) if e_i is not None and e_i < len(row) else ""
            if not email:
                continue
            label = row[b_i] if b_i is not None and b_i < len(row) else ""
            n = batch_num(label)
            if n is None:
                continue
            out[n].append({
                "email": email,
                "phone": _norm_phone(row[p_i]) if p_i is not None and p_i < len(row) else "",
                "refund": is_refunded(row[r_i]) if r_i is not None and r_i < len(row) else False,
                "label": re.sub(r"\s+", " ", str(label or "")).strip(),
            })
        added = sum(len(v) for v in out.values()) - n_before
        if not added:
            warnings.append(f'Tab "{ws.title}": no BSIAI batch rows - skipped')

    wb.close()
    return dict(out), warnings


# ──────────────────────────────── sessions ───────────────────────────────────
def sessions_from_files(attendee_files, l2_map) -> tuple[dict, list]:
    """Attendee files + the L2 map -> ({batch_num: {webinar: session}}, warnings).

    `attendee_files` is [(path, bytes), …] exactly as `live_data` yields it, and
    `l2_map` is `attendance_core.parse_l2`'s {webinar_id: (keys, topic)}.

    Two things this has to survive, both real:

    * **The same session exists on two Shared Drives** ("Weekly Sessions Files"
      as `BSI B1`, "Zoom extracts" as `BSIAI B1`) and occasionally twice within
      one drive. Keying on the webinar id collapses every copy; the copy with
      the most attendees wins, because the copies differ by a row or two and
      taking the larger never invents attendance.
    * **A file can parse to zero attendees** (Zoom sometimes exports a flat list
      with no `Attended`/`First Name` header). That is reported as a warning
      instead of silently marking a whole batch absent.
    """
    best: dict[int, dict] = defaultdict(dict)
    warnings: list[str] = []
    seen_zero: set[str] = set()

    for path, blob in attendee_files or ():
        fname = path.rsplit("/", 1)[-1]
        m = re.match(r"attendee_(\d+)_((?:20\d\d)_\d{2}_\d{2})", fname)
        if not m:
            continue
        wid, stamp = m.group(1), m.group(2)
        entry = l2_map.get(wid)
        if not entry:
            warnings.append(f"Webinar {wid} ({fname}): not in L2 - skipped")
            continue
        keys, topic = entry
        bsi = sorted({n for track, n in keys if track == "BSIAI"})
        if not bsi:
            continue

        text = blob.decode("utf-8-sig", errors="replace") if isinstance(blob, bytes) else blob
        emails, ph_full, ph10 = parse_attendees(text)
        if not emails and not ph_full:
            if wid not in seen_zero:
                seen_zero.add(wid)
                warnings.append(f"Webinar {wid} ({fname}): parsed to ZERO attendees "
                                "- not a Zoom Attendee Report? skipped")
            continue

        for n in bsi:
            prev = best[n].get(wid)
            if prev and len(prev["emails"]) >= len(emails):
                continue
            best[n][wid] = {
                "wid": wid,
                "mm": _mmdd(stamp),
                "topic": topic or f"Session {wid}",
                "emails": emails, "ph_full": ph_full, "ph10": ph10,
            }
    return dict(best), warnings


# ──────────────────────────────── assembly ───────────────────────────────────
# Engagement bands. BSIAI has no Payment/Close Type to slice by, so the question
# the panel answers is the one the data CAN answer: of the people who enrolled,
# how many actually keep turning up? A batch averaging 55% could be 55% of people
# attending everything or everyone attending half - those need different action.
# Half-open [lo, hi) so the labels mean what they say: a student who attended
# exactly half the course lands in "50-79%", not in "20-49%".
_BANDS = (("Attended 80%+",           lambda r: r >= 0.80),
          ("Most sessions (50-79%)",  lambda r: 0.50 <= r < 0.80),
          ("Some sessions (20-49%)",  lambda r: 0.20 <= r < 0.50),
          ("Barely (1-19%)",          lambda r: 0 < r < 0.20),
          ("Never attended",          lambda r: r == 0))


def engagement_bands(students: list, sessions: dict) -> list:
    """Per-student attendance rate -> the five bands, biggest first.

    Only sessions the student could attend count, and refunded students are left
    out entirely - they are not absentees, they left.
    """
    active = [s for s in students if not s["refund"]]
    n = len(sessions)
    if not active or not n:
        return []
    rates = []
    for st in active:
        p = st["phone"]
        hits = 0
        for sess in sessions.values():
            if (st["email"] and st["email"] in sess["emails"]) or                (p and (p in sess["ph_full"] or (len(p) >= 10 and p[-10:] in sess["ph10"]))):
                hits += 1
        rates.append(hits / n)
    out = []
    for label, in_band in _BANDS:
        grp = [r for r in rates if in_band(r)]
        if not grp:
            continue
        out.append({"type": label, "count": len(grp),
                    "pct": round(len(grp) / len(active) * 100, 1),
                    "att": round(sum(grp) / len(grp) * 100, 1)})
    return out


def build(roster: dict, sessions: dict, ratings: dict | None = None) -> tuple[dict, dict]:
    """({batch: [student]}, {batch: {wid: session}}) -> (DATA, summary).

    Same shape `data.build` returns so `dash_view.render` needs no special case.
    A student is Present when their email OR their phone appears in the attendee
    report; blank identities never match. Phones are compared on their last 10
    digits, which is what makes the +91/country-code variations line up.
    """
    DATA: dict[str, dict] = {}

    for n in sorted(set(roster) & set(sessions)):
        students = roster[n]
        strength = len(students)
        if not strength:
            continue
        active = sum(1 for s in students if not s["refund"])

        rows = []
        for sess in sessions[n].values():
            present = 0
            for st in students:
                if st["refund"]:
                    continue
                p = st["phone"]
                if (st["email"] and st["email"] in sess["emails"]) or \
                   (p and (p in sess["ph_full"] or (len(p) >= 10 and p[-10:] in sess["ph10"]))):
                    present += 1
            rate = (ratings or {}).get(sess["wid"]) or {}
            rows.append({
                "rating": rate.get("session"), "rating_trainer": rate.get("trainer"),
                "rating_recommend": rate.get("recommend"),
                "rating_n": rate.get("responses", 0),
                "col": None, "mm": sess["mm"],
                "date_lbl": date_label(sess["mm"]) or sess["wid"],
                "topic": sess["topic"],
                "l2_batch": f"BSIAI B{n}",
                "present": present, "absent": max(0, active - present),
                "total": active,
                "pct": round(present / active * 100, 1) if active else 0.0,
                "present_only": False, "no_l2": False,
            })

        if not rows:
            continue
        if all(r["mm"] for r in rows):
            rows.sort(key=lambda r: r["mm"])
        pcts = [r["pct"] for r in rows]
        DATA[f"B{n}"] = {
            "code": f"B{n}", "strength": strength, "active": active,
            "n_sessions": len(rows),
            "avg_pct": round(sum(pcts) / len(pcts), 1),
            "peak": max(pcts), "low": min(pcts),
            "sessions": rows,
            # The dashboard's generic panel renders whatever is here; for BSIAI
            # that is engagement bands rather than closing channels.
            "closing": engagement_bands(students, sessions[n]),
        }

    summary = {
        "batches": len(DATA),
        "enrolled": sum(d["strength"] for d in DATA.values()),
        "active": sum(d["active"] for d in DATA.values()),
        "sessions": sum(d["n_sessions"] for d in DATA.values()),
    }
    return DATA, summary


def analyse(roster_bytes: bytes, attendee_files, l2_map, ratings: dict | None = None):
    """One call: roster bytes + attendee files + L2 -> (DATA, summary, warnings)."""
    roster, w1 = read_roster(roster_bytes)
    sessions, w2 = sessions_from_files(attendee_files, l2_map)
    DATA, summary = build(roster, sessions, ratings)

    warnings = list(w1) + list(w2)
    for n in sorted(set(roster) - set(sessions)):
        warnings.append(f"BSIAI B{n}: {len(roster[n])} enrolled but no L2-known session yet")
    for n in sorted(set(sessions) - set(roster)):
        warnings.append(f"BSIAI B{n}: {len(sessions[n])} session(s) but no roster rows")
    return DATA, summary, warnings

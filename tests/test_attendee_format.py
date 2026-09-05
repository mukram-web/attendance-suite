"""The Zoom export-shape guard (stdlib unittest, no pytest).

Zoom emits two shapes for the same webinar. Only the Attendee Report is usable;
the flat participant list has no Attended column and no Phone, and used to parse
to zero attendees WITHOUT raising in both AI CAP paths. That marked a whole batch
absent, published a real-looking 0% day one, and left the weekly run green.

These tests pin the three things that must stay true:
  1. the two shapes are told apart,
  2. a GENUINELY empty report is still allowed to be empty,
  3. each of the three callers refuses the bad shape in its own way
     (marker skips, day-1 raises, BSIAI warns).
"""
import io
import os
import sys
import unittest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import attendance_core as ac  # noqa: E402
import bsiai  # noqa: E402
import day1_analysis  # noqa: E402


# ── the two shapes, as Zoom actually writes them ────────────────────────────
def report_csv(rows=(("a@x.com", "919000000001"), ("b@x.com", "919000000002")),
               sections=True):
    """A real Attendee Report. `sections=False` drops the named-section rows,
    which is how the BSIAI fixtures write it — still a valid report."""
    out = ["Attendee Report",
           "Report generated time,09/09/2026 05:29:06 AM",
           "",
           "Topic,Webinar ID,Actual Start Time,Unique Viewers",
           f"Some Session,91695866411,09/09/2026 19:00,{len(rows)}",
           ""]
    if sections:
        out.append("Attendee Details")
    out.append("Attended,First Name,Last Name,Email,Phone,Join Time,Leave Time,"
               "Time in Session (minutes)")
    for email, phone in rows:
        out.append(f"Yes,First,Last,{email},{phone},09/09/2026 19:02:00,"
                   f"09/09/2026 20:31:00,89")
    return "\n".join(out)


FLAT = ("Name (original name),Email,Join time,Leave time,Duration (minutes),"
        "Guest,Recording disclaimer response,In waiting room\n"
        "Test One,a@x.com,09/09/2026 19:02,09/09/2026 20:31,89,Yes,,No\n"
        "Test Two,b@x.com,09/09/2026 19:05,09/09/2026 20:30,85,Yes,,No\n")


def tab_report():
    """A REAL Attendee Report that is tab-delimited instead of comma.

    This is the case shape-detection alone gets wrong: the "Attendee Details"
    line survives any delimiter, so the file still looks like a report, but
    parse_attendees reads the comma dialect and gets nobody out of it. It must
    still be refused, and for the delimiter reason rather than the flat-list one.
    """
    return report_csv().replace(",", "\t")


class TestIsAttendeeReport(unittest.TestCase):
    def test_the_flat_participant_list_is_rejected(self):
        self.assertFalse(ac.is_attendee_report(FLAT))

    def test_a_real_report_is_accepted(self):
        self.assertTrue(ac.is_attendee_report(report_csv()))

    def test_a_report_without_section_rows_is_still_accepted(self):
        # The Attended/First Name header alone is enough — this is the shape the
        # BSIAI fixtures use, and rejecting it would break real marking.
        self.assertTrue(ac.is_attendee_report(report_csv(sections=False)))

    def test_a_genuinely_empty_report_is_NOT_a_format_error(self):
        # Header present, no data rows: nobody joined. That is a real zero and
        # must stay allowed, or a dead session would fail the whole week.
        empty = report_csv(rows=())
        self.assertTrue(ac.is_attendee_report(empty))
        self.assertEqual(ac.parse_attendees(empty), (set(), set(), set()))

    def test_empty_and_junk_input_are_rejected(self):
        for bad in ("", "   ", "not,a,csv,at,all\n1,2,3,4,5\n"):
            self.assertFalse(ac.is_attendee_report(bad), repr(bad))

    def test_the_tag_appears_in_the_shared_constant(self):
        # pipeline.py greps warnings for this string; keep them in one place.
        self.assertIn("Attendee Report", ac.UNREADABLE_FORMAT_TAG)
        self.assertIn("ZERO", ac.ZERO_ATTENDEE_TAG)

    def test_a_tab_delimited_report_still_LOOKS_like_a_report(self):
        # Shape detection alone cannot catch this one — which is exactly why the
        # callers gate on the parse RESULT, not on this predicate.
        self.assertTrue(ac.is_attendee_report(tab_report()))
        self.assertEqual(ac.parse_attendees(tab_report()), (set(), set(), set()))


class TestZeroAttendeeReason(unittest.TestCase):
    def test_the_flat_list_is_named_as_a_shape_problem(self):
        why = ac.zero_attendee_reason(FLAT)
        self.assertIn(ac.UNREADABLE_FORMAT_TAG, why)

    def test_a_tab_delimited_report_is_named_as_a_delimiter_problem(self):
        why = ac.zero_attendee_reason(tab_report())
        self.assertNotIn(ac.UNREADABLE_FORMAT_TAG, why)
        self.assertIn("delimiter", why)


class TestParseAttendeesUnchanged(unittest.TestCase):
    """The guard must not change what a GOOD file parses to."""

    def test_a_real_report_still_parses_both_keys(self):
        em, pf, p10 = ac.parse_attendees(report_csv())
        self.assertEqual(em, {"a@x.com", "b@x.com"})
        self.assertEqual(pf, {"919000000001", "919000000002"})
        self.assertEqual(p10, {"9000000001", "9000000002"})

    def test_the_flat_list_still_parses_to_zero(self):
        # Unchanged behaviour: parse_attendees stays total and returns empties.
        # It is the CALLERS that now refuse to act on that.
        self.assertEqual(ac.parse_attendees(FLAT), (set(), set(), set()))


class TestDay1Raises(unittest.TestCase):
    def test_read_attendees_raises_on_the_flat_list(self):
        with self.assertRaises(ValueError) as cm:
            day1_analysis.read_attendees(FLAT.encode())
        self.assertIn(ac.UNREADABLE_FORMAT_TAG, str(cm.exception))

    def test_read_attendees_raises_on_a_tab_delimited_report(self):
        # The case shape-detection missed: refused because nothing parsed.
        with self.assertRaises(ValueError) as cm:
            day1_analysis.read_attendees(tab_report().encode())
        self.assertIn(ac.ZERO_ATTENDEE_TAG, str(cm.exception))

    def test_read_attendees_raises_on_a_report_with_no_rows(self):
        # A day one where not one row is readable is not a real 0%.
        with self.assertRaises(ValueError):
            day1_analysis.read_attendees(report_csv(rows=()).encode())

    def test_read_attendees_still_reads_a_real_report(self):
        people, viewers = day1_analysis.read_attendees(report_csv().encode())
        self.assertEqual(set(people), {"a@x.com", "b@x.com"})
        self.assertEqual(viewers, 2)

    def test_an_all_internal_report_returns_empty_WITHOUT_raising(self):
        # people is non-empty, attended is empty. That is the internal-account
        # filter working, not a broken file, so it must not raise.
        raw = report_csv(rows=(("staff@be10x.com", "919000000009"),)).encode()
        attended, _ = day1_analysis.read_attendees(raw)
        self.assertEqual(attended, {})


class TestMarkerSkips(unittest.TestCase):
    """The marker must SKIP the session, not mark the batch absent."""

    def _roster(self):
        from openpyxl import Workbook
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("AI CAP B17")
        ws.append(["Country", "Registered Number", "Registered Mail", "WhatsApp",
                   "Broadcast", "Batch", "Amount", "Payment", "Close Type", "POD"])
        ws.append([91, "919000000001", "a@x.com", "", "", "B17", 0,
                   "Full Paid", "BDA Closing", ""])
        ws.append([91, "919000000002", "b@x.com", "", "", "B17", 0,
                   "Full Paid", "BDA Closing", ""])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _l2(self):
        from openpyxl import Workbook
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("Sep 2026")
        ws.append(["Date", "Webinar ID", "Batch Name", "Topic"])
        ws.append(["09/09/2026", "91695866411", "AI CAP B17", "Some Session"])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _run(self, payload):
        name = "2026-09-09 - AI CAP B17 - Some Session/attendee_91695866411_2026_09_09.csv"
        return ac.process_files(self._roster(), self._l2(),
                                [(name, payload.encode())], values_only=True)

    def test_a_flat_file_marks_nothing_and_warns(self):
        _, report, warnings = self._run(FLAT)
        self.assertEqual(report, [], "a bad export must create NO session column")
        self.assertTrue(any(ac.UNREADABLE_FORMAT_TAG in w for w in warnings),
                        f"expected a format warning, got {warnings}")
        self.assertTrue(any("SKIPPED" in w for w in warnings))

    def test_a_tab_delimited_report_is_also_skipped(self):
        _, report, warnings = self._run(tab_report())
        self.assertEqual(report, [], "an unparseable report must mark nobody")
        self.assertTrue(any(ac.ZERO_ATTENDEE_TAG in w for w in warnings))

    def test_the_warning_carries_the_tag_the_pipeline_gate_greps(self):
        # pipeline.py refuses to publish on this exact substring. If the message
        # is reworded without the constant, the gate silently stops firing.
        _, _, warnings = self._run(FLAT)
        self.assertTrue(any(ac.ZERO_ATTENDEE_TAG in w for w in warnings),
                        f"gate would not fire on: {warnings}")

    def test_a_good_file_still_marks_the_session(self):
        _, report, _ = self._run(report_csv())
        self.assertEqual(len(report), 1, "a good export must still be marked")
        self.assertEqual(report[0]["present"], 2)


class TestBsiaiStillWarns(unittest.TestCase):
    L2 = {"111": (frozenset({("BSIAI", 1)}), "WhatsApp")}

    def test_the_flat_list_is_skipped_and_named_as_a_format_problem(self):
        files = [("2026-09-09 - BSI B1 - X/attendee_111_2026_09_09.csv",
                  FLAT.encode())]
        sess, warns = bsiai.sessions_from_files(files, self.L2)
        self.assertEqual(sess, {})
        self.assertTrue(any("ZERO" in w for w in warns))
        self.assertTrue(any(ac.UNREADABLE_FORMAT_TAG in w for w in warns),
                        f"the warning should name the cause, got {warns}")

    def test_a_real_report_is_still_accepted(self):
        files = [("2026-09-09 - BSI B1 - X/attendee_111_2026_09_09.csv",
                  report_csv().encode())]
        sess, _ = bsiai.sessions_from_files(files, self.L2)
        self.assertEqual(len(sess[1]["111"]["emails"]), 2)



class TestDuplicateCopiesAcrossDrives(unittest.TestCase):
    """The same webinar sits on BOTH Shared Drives (166 sessions as of Sep 2026),
    and the two use identical folder and file names, so the copies arrive as two
    entries with the same path. Only one is marked; these pin which."""

    # Same path twice, exactly as the two drives produce it.
    PATH = "2026-09-09 - AI CAP B17 - Some Session/attendee_91695866411_2026_09_09.csv"

    BOTH = (("a@x.com", "919000000001"), ("b@x.com", "919000000002"))
    ONE = (("a@x.com", "919000000001"),)

    def _run(self, payloads):
        """payloads are listed in attendee_folder_id order: Weekly Sessions
        first, Zoom extracts last."""
        marker = TestMarkerSkips()
        files = [(self.PATH, p.encode()) for p in payloads]
        return ac.process_files(marker._roster(), marker._l2(), files,
                                values_only=True)

    def test_the_LAST_listed_drive_wins_when_both_copies_are_good(self):
        # Zoom extracts is listed last and is authoritative (owner, 2026-09-06).
        _, report, _ = self._run([report_csv(self.ONE), report_csv(self.BOTH)])
        self.assertEqual(report[0]["present"], 2, "the last copy should have won")

    def test_the_preference_is_an_order_rule_not_a_bigger_file_rule(self):
        # Reversed input must flip the winner — otherwise something else (size,
        # dict ordering) is deciding and swapping the secret would not be safe.
        _, report, _ = self._run([report_csv(self.BOTH), report_csv(self.ONE)])
        self.assertEqual(report[0]["present"], 1, "the last copy should have won")

    def test_a_good_copy_rescues_an_unreadable_preferred_copy(self):
        # The whole point: a flat export on the preferred drive must not cost a
        # session the other drive holds intact.
        _, report, warnings = self._run([report_csv(self.BOTH), FLAT])
        self.assertEqual(len(report), 1, "the readable copy should have marked it")
        self.assertEqual(report[0]["present"], 2)
        self.assertFalse([w for w in warnings if ac.ZERO_ATTENDEE_TAG in w],
                         f"the publish gate must not fire: {warnings}")

    def test_the_rescue_works_in_either_direction(self):
        _, report, warnings = self._run([FLAT, report_csv(self.BOTH)])
        self.assertEqual(report[0]["present"], 2)
        self.assertFalse([w for w in warnings if ac.ZERO_ATTENDEE_TAG in w])

    def test_all_copies_unreadable_still_skips_and_trips_the_gate(self):
        _, report, warnings = self._run([FLAT, tab_report()])
        self.assertEqual(report, [])
        gate = [w for w in warnings if ac.ZERO_ATTENDEE_TAG in w]
        self.assertTrue(gate, f"the gate must still fire: {warnings}")
        self.assertIn("all 2 copy(ies)", gate[0])
        # Both causes named, so the fix is obvious without re-downloading.
        self.assertIn(ac.UNREADABLE_FORMAT_TAG, gate[0])
        self.assertIn("delimiter", gate[0])

    def test_a_numbered_duplicate_never_beats_a_real_copy(self):
        # "attendee_..(1).csv" is Drive's re-upload artefact, not a newer export.
        marker = TestMarkerSkips()
        dup = self.PATH.replace(".csv", " (1).csv")
        files = [(self.PATH, report_csv(self.BOTH).encode()),
                 (dup, report_csv(self.ONE).encode())]
        _, report, _ = ac.process_files(marker._roster(), marker._l2(), files,
                                        values_only=True)
        self.assertEqual(report[0]["present"], 2, "the (1) copy must not win")



class TestEitherKeyIsEnough(unittest.TestCase):
    """A report needs Email OR Phone, not both. Matching is an OR, so one column
    still marks every student whose roster row agrees on it (owner, 2026-09-06)."""

    def _rep(self, header, rows):
        return chr(10).join(["Attendee Report", "", header] + rows)

    BOTH = ("Attended,First Name,Last Name,Email,Phone,Join Time",
            ["Yes,A,B,a@x.com,919000000001,x", "Yes,C,D,b@x.com,919000000002,x"])
    EMAIL_ONLY = ("Attended,First Name,Last Name,Email,Join Time",
                  ["Yes,A,B,a@x.com,x", "Yes,C,D,b@x.com,x"])
    PHONE_ONLY = ("Attended,First Name,Last Name,Phone,Join Time",
                  ["Yes,A,B,919000000001,x", "Yes,C,D,919000000002,x"])

    def test_email_and_phone(self):
        em, pf, p10 = ac.parse_attendees(self._rep(*self.BOTH))
        self.assertEqual(em, {"a@x.com", "b@x.com"})
        self.assertEqual(pf, {"919000000001", "919000000002"})

    def test_email_only_column(self):
        em, pf, _ = ac.parse_attendees(self._rep(*self.EMAIL_ONLY))
        self.assertEqual(em, {"a@x.com", "b@x.com"}, "emails must survive a missing Phone column")
        self.assertEqual(pf, set())

    def test_phone_only_column(self):
        em, pf, p10 = ac.parse_attendees(self._rep(*self.PHONE_ONLY))
        self.assertEqual(em, set())
        self.assertEqual(pf, {"919000000001", "919000000002"})
        self.assertEqual(p10, {"9000000001", "9000000002"})

    def test_blank_phone_values_still_yield_emails(self):
        em, pf, _ = ac.parse_attendees(self._rep(
            "Attended,First Name,Last Name,Email,Phone,Join Time",
            ["Yes,A,B,a@x.com,,x", "Yes,C,D,b@x.com,,x"]))
        self.assertEqual(em, {"a@x.com", "b@x.com"})
        self.assertEqual(pf, set())

    def test_neither_column_is_still_unusable(self):
        em, pf, _ = ac.parse_attendees(self._rep(
            "Attended,First Name,Last Name,Join Time", ["Yes,A,B,x"]))
        self.assertEqual((em, pf), (set(), set()))

    def test_short_rows_do_not_crash_either_shape(self):
        # A trailing/ragged row must be skipped, not raise.
        for header, rows in (self.EMAIL_ONLY, self.PHONE_ONLY, self.BOTH):
            ac.parse_attendees(self._rep(header, list(rows) + ["Yes", "", "Yes,A"]))

    def test_the_marker_marks_from_an_email_only_report(self):
        marker = TestMarkerSkips()
        _, report, warnings = marker._run(self._rep(*self.EMAIL_ONLY))
        self.assertEqual(len(report), 1, f"session should be marked: {warnings}")
        self.assertEqual(report[0]["present"], 2)
        self.assertFalse([w for w in warnings if ac.ZERO_ATTENDEE_TAG in w])

    def test_the_marker_marks_from_a_phone_only_report(self):
        marker = TestMarkerSkips()
        _, report, warnings = marker._run(self._rep(*self.PHONE_ONLY))
        self.assertEqual(len(report), 1, f"session should be marked: {warnings}")
        self.assertEqual(report[0]["present"], 2)
        self.assertFalse([w for w in warnings if ac.ZERO_ATTENDEE_TAG in w])


if __name__ == "__main__":
    unittest.main()

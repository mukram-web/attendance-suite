"""
pipeline.py — the weekly heavy job, run OFF the Streamlit server.

Fetches the roster / L2 / new Zoom attendee reports from Google Drive, marks
attendance, builds every artefact the dashboard renders, packs them into ONE
small DuckDB file, and uploads it (plus the marked .xlsx) to a private Drive
folder. The deployed app then just downloads that file and renders — no
fetching, no marking, no crashes on Streamlit Cloud's small instance.

Runs on GitHub Actions every Monday 06:00 IST (see .github/workflows/refresh.yml)
or manually:

    python pipeline.py                # full run: fetch → mark → build → upload
    python pipeline.py --no-upload    # build .cache/attendance.duckdb only

Credentials & config resolve in this order:
  1. env vars  GDRIVE_SERVICE_ACCOUNT_JSON (the key JSON, one line)
               ROSTER_ID / L2_ID / ATTENDEE_FOLDER_ID / STORE_FOLDER_ID
  2. local fallback: .streamlit/secrets.toml + the service-account .json key
     file sitting next to this script (developer machine).

The store schema (attendance.duckdb):
  meta(key, value)       — JSON blobs: DATA, summary, report, warnings, source,
                           generated_at, batches, sheet_map, marked_xlsx_file_id,
                           day1 (the Day-1 analysis for the newest batches —
                           aggregates only), forecast (predicted attendance for
                           sessions that have not run yet — aggregates only)
  compute                — dashboard_core.compute() table (per batch × session)
  grid_<batch>           — roster_grid() per batch (the Roster tab, incl. PII —
                           the store must stay in a PRIVATE Drive folder)
"""
from __future__ import annotations

import argparse
import glob
import io
import json
import os
import re
import sys
from datetime import datetime, timedelta, timezone

import duckdb
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)

import attendance_core as ac          # noqa: E402
import bsiai                          # noqa: E402
import dashboard_core as dc           # noqa: E402
import data as ddata                  # noqa: E402
import polls                          # noqa: E402
import day1_analysis                  # noqa: E402
import forecast                       # noqa: E402
import archive                        # noqa: E402
import sheets as dsheets              # noqa: E402
import live_data                      # noqa: E402

# How many of the newest batches the day-1 dashboard covers. The window rolls by
# itself: a new batch tab joins, the oldest drops off.
DAY1_BATCHES = 4

# How far ahead the forecast runs. Eight weeks is where the backtest still holds
# under ~13% MAPE and the curriculum sheet is actually filled in; past that the
# schedule thins out and the error climbs, so forecasting further would be
# confident-looking noise.
FORECAST_WEEKS = 8

IST = timezone(timedelta(hours=5, minutes=30))
STORE_NAME = "attendance.duckdb"
MARKED_NAME = "Master_Batch_Rosters_marked.xlsx"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
RW_SCOPES = ["https://www.googleapis.com/auth/drive"]


# ─────────────────────────── config & credentials ────────────────────────────
def _secrets_toml() -> dict:
    """Minimal read of .streamlit/secrets.toml for local runs (no streamlit)."""
    path = os.path.join(HERE, ".streamlit", "secrets.toml")
    if not os.path.exists(path):
        return {}
    try:
        import tomllib
        with open(path, "rb") as fh:
            return tomllib.load(fh)
    except Exception:
        return {}


def load_config() -> dict:
    toml = _secrets_toml()
    drive = toml.get("drive", {})

    def pick(env_key, toml_key):
        return os.environ.get(env_key) or drive.get(toml_key) or ""

    cfg = {
        "roster_id": pick("ROSTER_ID", "roster_id"),
        "l2_id": pick("L2_ID", "l2_id"),
        "attendee_folder_id": pick("ATTENDEE_FOLDER_ID", "attendee_folder_id"),
        "store_folder_id": pick("STORE_FOLDER_ID", "store_folder_id"),
        # Optional: the BSIAI programme's own roster Sheet. Absent -> the BSIAI
        # section is simply not built, and the app shows its tab as unconfigured.
        # Like every other id, the Sheet must be shared with the service account.
        "bsiai_roster_id": pick("BSIAI_ROSTER_ID", "bsiai_roster_id"),
        # Optional: the Master Curriculum Schedule Sheet — what is PLANNED, one
        # tab per pod. Absent -> no forecast section and the tab says so. Note
        # this Sheet is owned outside the team that owns the roster, so sharing
        # it with the service account is a separate step that is easy to forget.
        "curriculum_id": pick("CURRICULUM_ID", "curriculum_id"),
    }

    env_json = os.environ.get("GDRIVE_SERVICE_ACCOUNT_JSON")
    if env_json:
        cfg["sa_info"] = json.loads(env_json)
    elif toml.get("gcp_service_account"):
        cfg["sa_info"] = dict(toml["gcp_service_account"])
    else:
        # Match on the FILE NAME, never the full path — a checkout under e.g.
        # .../customer-service/ would otherwise match every json in the folder
        # and hand a data file to the credential parser.
        keys = [p for p in glob.glob(os.path.join(HERE, "*.json"))
                if _looks_like_sa_key(p)
                or "service" in os.path.basename(p).lower()
                or "fresh-delight" in os.path.basename(p).lower()]
        if not keys:
            raise SystemExit("No service-account credentials: set "
                             "GDRIVE_SERVICE_ACCOUNT_JSON or add the key file.")
        with open(keys[0], encoding="utf-8") as fh:
            cfg["sa_info"] = json.load(fh)

    missing = [k for k in ("roster_id", "attendee_folder_id") if not cfg[k]]
    if missing:
        raise SystemExit(f"Missing config: {', '.join(missing)} "
                         "(env vars or .streamlit/secrets.toml [drive])")
    return cfg


def _looks_like_sa_key(path: str) -> bool:
    try:
        with open(path, encoding="utf-8") as fh:
            head = fh.read(2048)
        return '"type": "service_account"' in head
    except OSError:
        return False


# ───────────────────────────── build the store ───────────────────────────────
def _prepend_intro_sessions(DATA: dict) -> None:
    """Same rule as the app: first session of every batch = the intro call,
    counts from intro_attendance.json (aggregates precomputed from L2 customer)."""
    path = os.path.join(HERE, "intro_attendance.json")
    if not os.path.exists(path):
        # Silence here would mean a store whose batches quietly lack their intro
        # row — most likely the file wasn't committed (it's under a *.json ignore).
        print("   WARNING: intro_attendance.json is missing — the store will have "
              "no 'Intro call' session. Commit it alongside pipeline.py.",
              flush=True)
        return
    with open(path, encoding="utf-8") as fh:
        intro = json.load(fh)
    for code, d in DATA.items():
        att = intro.get(code)
        if not att:
            continue
        stg = d["strength"]
        d["sessions"].insert(0, {
            "col": None, "mm": None,
            "date_lbl": "Intro call",
            "topic": "Intro call (pre-batch)",
            "present": att, "absent": max(0, stg - att), "total": stg,
            "pct": round(min(100.0, 100 * att / stg), 1),
            "present_only": False, "no_l2": False, "is_intro": True,
            # the intro call predates the batch and has no feedback poll
            "rating": None, "rating_n": 0,
        })


def build_store(path: str, marked_bytes: bytes, report, warnings, source: str,
                l2_bytes, attendee_names, marked_xlsx_file_id: str,
                stamps: dict, day1: dict | None = None,
                bsiai_section: dict | None = None,
                ratings: dict | None = None,
                curric_tabs: dict | None = None) -> dict:
    """Write attendance.duckdb next to nothing else — one self-contained file."""
    # dashboard DATA/summary (same code path the app used to run at startup)
    wb = load_workbook(io.BytesIO(marked_bytes), read_only=True, data_only=True)
    tabs = {ws.title: [list(r) for r in ws.iter_rows(values_only=True)]
            for ws in wb.worksheets}
    wb.close()
    topics, l2_labels, mentors = dsheets.webinar_topic_lookup(
        [(n, None) for n in attendee_names], l2_bytes,
        with_labels=True, with_mentors=True)
    DATA, summary = ddata.build(tabs, topics, l2_labels, ratings, mentors)
    _prepend_intro_sessions(DATA)

    # Forecast for sessions that have not run yet. Built HERE because it needs the
    # finished DATA and nothing else — no extra fetch, and the numbers can never
    # disagree with the dashboard they sit beside. None means "not configured";
    # a section carrying only warnings means "configured, could not build".
    forecast_section = None
    if curric_tabs is not None:
        try:
            forecast_section = forecast.build(DATA, curric_tabs,
                                              datetime.now(IST).date(),
                                              horizon_weeks=FORECAST_WEEKS)
        except Exception as e:
            forecast_section = forecast.empty_result(f"forecast build failed: {e}")

    df = dc.compute(marked_bytes)
    smap = dc.batch_sheet_map(marked_bytes)
    # only real batch tabs — helper tabs like 'l2 cx data' or 'Auto pay' pass
    # dashboard_core's loose sheet filter but are not batches
    batches = sorted((b for b in df["Batch"].unique()
                      if b in smap and re.fullmatch(r"B\d+", str(b))),
                     key=dc.batch_key)
    df = df[df["Batch"].isin(batches)].reset_index(drop=True)

    if os.path.exists(path):
        os.remove(path)
    con = duckdb.connect(path)
    con.register("df_compute", df)
    con.execute("CREATE TABLE compute AS SELECT * FROM df_compute")

    grid_rows = {}
    for b in batches:
        grid = dc.roster_grid(marked_bytes, smap[b])
        grid_rows[b] = len(grid)
        con.register("df_grid", grid)
        con.execute(f'CREATE TABLE "grid_{b}" AS SELECT * FROM df_grid')
        con.unregister("df_grid")

    meta = {
        "generated_at": datetime.now(IST).strftime("%d %b %Y, %H:%M IST"),
        "generated_at_iso": datetime.now(IST).isoformat(),
        "DATA": DATA,
        "summary": summary,
        "report": report,
        "warnings": warnings,
        "source": source,
        "batches": batches,
        "sheet_map": smap,
        "marked_xlsx_file_id": marked_xlsx_file_id,
        "stamps": stamps,
        "day1": day1 or {"batches": [], "skipped": []},
        # None = no CURRICULUM_ID configured. A section present but with empty
        # `sessions` means it was configured and produced nothing — the app says
        # which, because "not set up" and "set up but broken" need different fixes.
        "forecast": forecast_section,
        # None means "not configured" (no BSIAI_ROSTER_ID). A section whose DATA
        # is empty means "configured, nothing to show yet". The app distinguishes.
        "bsiai": bsiai_section,
    }
    con.execute("CREATE TABLE meta (key VARCHAR, value VARCHAR)")
    con.executemany("INSERT INTO meta VALUES (?, ?)",
                    [(k, json.dumps(v)) for k, v in meta.items()])
    con.close()
    return {"batches": len(batches), "students": sum(grid_rows.values()),
            "sessions": summary["sessions"]}


# ────────────────────────────────── main ─────────────────────────────────────
def main() -> None:
    # Windows consoles default to cp1252, which can't print every character in
    # Drive file names — never let logging kill the pipeline.
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--no-upload", action="store_true",
                    help="build the store locally, skip the Drive upload")
    ap.add_argument("--mode", default="exact", choices=["exact", "inclusive"],
                    help="marker matching rule (default: exact, same as the app)")
    ap.add_argument("--no-site", action="store_true",
                    help="skip rendering site/ (the static website)")
    ap.add_argument("--allow-partial", action="store_true",
                    help="publish even if some attendee files/folders failed "
                         "(default: refuse, so a complete store is never "
                         "overwritten by a partial one)")
    args = ap.parse_args()

    cfg = load_config()
    live_data.set_service_account(cfg["sa_info"], scopes=RW_SCOPES)
    svc = live_data._drive_service()

    print("[1/8] Fetching roster + L2 …", flush=True)
    roster_bytes, roster_stamp = live_data.fetch_sheet_cached(svc, cfg["roster_id"])
    l2_bytes, l2_stamp = (live_data.fetch_sheet_cached(svc, cfg["l2_id"])
                          if cfg["l2_id"] else (None, ""))
    print(f"   roster {len(roster_bytes):,} bytes (modified {roster_stamp})")

    print("[2/8] Fetching attendee reports (new sessions only) …", flush=True)
    attendee_files, info = live_data.fetch_new_attendees(
        svc, cfg["attendee_folder_id"], roster_bytes)
    print(f"   {info['files']} file(s) from {info['new_folders']} folder(s), "
          f"{info['failed']} failed, {info['skipped_no_sheet']} without a roster tab")

    # Publishing a store built from partial data would overwrite a COMPLETE one
    # and the job would still go green — so refuse. Re-running picks up whatever
    # failed (successful downloads stay cached, so a retry is cheap).
    problems = []
    if info.get("listing_errors"):
        problems.append(f"{len(info['listing_errors'])} session folder(s) could not "
                        f"be listed: {'; '.join(info['listing_errors'][:5])}")
    if info.get("bad_zips"):
        problems.append(f"{len(info['bad_zips'])} unreadable zip(s): "
                        f"{'; '.join(info['bad_zips'][:5])}")
    if info.get("failed"):
        problems.append(f"{info['failed']} attendee file(s) failed to download: "
                        f"{'; '.join(info.get('download_errors', [])[:5])}")
    if problems and not args.allow_partial:
        raise SystemExit("Refusing to publish partial data — the previous store is "
                         "left in place:\n  - " + "\n  - ".join(problems)
                         + "\nRe-run the job; pass --allow-partial to override.")

    print("[3/8] Marking attendance …", flush=True)
    if attendee_files:
        marked_bytes, report, warnings = ac.process_files(
            roster_bytes, l2_bytes, attendee_files, mode=args.mode,
            values_only=True)
    else:
        marked_bytes, report, warnings = roster_bytes, [], []
    new_n = sum(1 for r in report if r["kind"] == "NEW")
    print(f"   {len(report)} session column(s) marked "
          f"({new_n} new) · {len(warnings)} warning(s)")

    if not args.no_upload and not cfg["store_folder_id"]:
        raise SystemExit("STORE_FOLDER_ID missing — where should the store "
                         "be uploaded? (or run with --no-upload)")

    # Day-1 analysis runs BEFORE anything is uploaded: if it refuses below, this
    # run must have left Drive exactly as it found it.
    print(f"[4/8] Day-1 analysis (newest {DAY1_BATCHES} batches) …", flush=True)
    day1 = {"batches": [], "skipped": [], "errors": []}
    try:
        wb_r = load_workbook(io.BytesIO(marked_bytes), read_only=True, data_only=True)
        roster_tabs = {ws.title: [list(r) for r in ws.iter_rows(values_only=True)]
                       for ws in wb_r.worksheets
                       if re.fullmatch(r"\s*AI\s*CAP\s*B\d+\s*", ws.title, re.I)}
        wb_r.close()
        day1 = day1_analysis.build(svc, roster_tabs, l2_bytes,
                                   cfg["attendee_folder_id"],
                                   n_batches=DAY1_BATCHES)
    except Exception as e:
        day1 = {"batches": [], "skipped": [], "errors": [f"analysis failed: {e}"]}

    # Same rule as the attendee gate above: a transiently broken day-1 analysis
    # must not overwrite last week's complete tab with an empty one and go green.
    if (day1.get("errors") or not day1["batches"]) and not args.allow_partial:
        raise SystemExit(
            "Refusing to publish: the day-1 analysis is incomplete — the previous "
            "store is left in place:\n  - "
            + "\n  - ".join(day1.get("errors") or ["no batches produced"])
            + "\nRe-run the job; pass --allow-partial to publish without it.")

    marked_xlsx_file_id = ""
    if not args.no_upload:
        print("[5/8] Uploading marked roster xlsx …", flush=True)
        marked_xlsx_file_id = live_data.upload_to_folder(
            svc, cfg["store_folder_id"], MARKED_NAME, marked_bytes, XLSX_MIME)
    elif cfg["store_folder_id"]:
        # --no-upload: keep pointing at the xlsx a previous full run left on
        # Drive, so a locally built store doesn't lose the download button if it
        # is copied into the store folder by hand.
        try:
            prev = live_data.find_in_folder(svc, cfg["store_folder_id"], MARKED_NAME)
            marked_xlsx_file_id = prev["id"] if prev else ""
        except Exception:
            marked_xlsx_file_id = ""

    # BSIAI — a separate programme with its own roster Sheet, its own batch
    # numbering and no session columns anywhere, so it is computed from the
    # attendee reports rather than marked into a workbook. Never fatal: a BSIAI
    # problem must not cost the AI CAP refresh its weekly publish.
    # Session feedback polls: one small CSV per session, disk-cached like the
    # attendee reports. Never fatal - a missing poll costs that session its
    # rating, not the run.
    print("[5a] Session feedback polls …", flush=True)
    ratings, ratings_by_wid = {}, {}
    try:
        poll_files, pinfo = live_data.fetch_polls(svc, cfg["attendee_folder_id"])
        ratings, ratings_by_wid = polls.lookup_by_session(poll_files, l2_bytes)
        print(f"   {pinfo['files']}/{pinfo['found']} poll file(s) · "
              f"{len(ratings_by_wid)} webinar(s) rated · "
              f"{len(ratings)} session key(s)")
    except Exception as e:
        print(f"   WARNING: poll ratings unavailable ({e}) - sessions show no rating.")

    bsiai_section = None
    if cfg["bsiai_roster_id"]:
        print("[5b] BSIAI attendance …", flush=True)
        try:
            b_roster, b_stamp = live_data.fetch_sheet_cached(svc, cfg["bsiai_roster_id"])
            b_files, b_info = live_data.fetch_track_attendees(
                svc, cfg["attendee_folder_id"], "BSIAI")
            l2_map = ac.parse_l2(l2_bytes) if l2_bytes else {}
            B_DATA, b_summary, b_warn = bsiai.analyse(b_roster, b_files, l2_map,
                                                      ratings_by_wid)
            bsiai_section = {"DATA": B_DATA, "summary": b_summary,
                             "warnings": b_warn, "stamp": b_stamp,
                             "source": (f"BSIAI roster + {b_info['files']} attendee "
                                        f"file(s) from {b_info['new_folders']} folder(s)")}
            print(f"   {b_summary['batches']} batch(es) · "
                  f"{b_summary['enrolled']:,} enrolled · "
                  f"{b_summary['sessions']} session(s) · {len(b_warn)} warning(s)")
        except Exception as e:
            bsiai_section = {"DATA": {}, "summary": {"batches": 0, "enrolled": 0,
                                                     "active": 0, "sessions": 0},
                             "warnings": [f"BSIAI build failed: {e}"],
                             "stamp": "", "source": ""}
            print(f"   WARNING: BSIAI section failed ({e}) - the tab will say so.")
    else:
        print("[5b] BSIAI skipped (no BSIAI_ROSTER_ID configured)", flush=True)

    # The Master Curriculum Schedule — what is PLANNED, one tab per pod. Read as
    # xlsx like every other Sheet so the same modifiedTime cache applies.
    # A failure here must never sink the run: the forecast is an extra, and last
    # week's dashboard is worth far more than a red build over a missing schedule.
    curric_tabs = None
    if cfg["curriculum_id"]:
        print("[5c] Curriculum schedule …", flush=True)
        try:
            c_bytes, _c_stamp = live_data.fetch_sheet_cached(svc, cfg["curriculum_id"])
            cwb = load_workbook(io.BytesIO(c_bytes), read_only=True, data_only=True)
            curric_tabs = {ws.title: [list(r) for r in ws.iter_rows(values_only=True)]
                           for ws in cwb.worksheets}
            cwb.close()
            print(f"   {len(curric_tabs)} tab(s)")
        except Exception as e:
            curric_tabs = {}          # {} = configured but unreadable; None = unset
            print(f"   WARNING: curriculum unavailable ({e}) - the Forecast tab "
                  "will say so.")
    else:
        print("[5c] Forecast skipped (no CURRICULUM_ID configured)", flush=True)

    print("[6/8] Building the DuckDB store …", flush=True)
    names = live_data.list_attendee_names(cfg["attendee_folder_id"])
    source = (f"Google Drive — roster, {info['files']} file(s) from "
              f"{info['new_folders']} new session(s)")
    os.makedirs(os.path.join(HERE, ".cache"), exist_ok=True)
    store_path = os.path.join(HERE, ".cache", STORE_NAME)
    stats = build_store(store_path, marked_bytes, report, warnings, source,
                        l2_bytes, names, marked_xlsx_file_id,
                        {"roster": roster_stamp, "l2": l2_stamp, "mode": args.mode},
                        day1=day1, bsiai_section=bsiai_section, ratings=ratings,
                        curric_tabs=curric_tabs)
    size = os.path.getsize(store_path)
    print(f"   {stats['batches']} batches · {stats['students']:,} students · "
          f"{stats['sessions']} sessions · {size / 1e6:.1f} MB")

    # The static website the team actually opens. Built from the store that was
    # just written, so the site can never disagree with the app.
    if not args.no_site:
        print("[7/8] Rendering the static site …", flush=True)
        import site_build
        site_build.build_site(store_path, os.path.join(HERE, "site"))

    if args.no_upload:
        print(f"[8/8] Skipped upload (--no-upload). Store at: {store_path}")
        return

    print("[8/8] Uploading the store …", flush=True)
    with open(store_path, "rb") as fh:
        store_bytes = fh.read()
    live_data.upload_to_folder(svc, cfg["store_folder_id"], STORE_NAME,
                               store_bytes)

    # Dated snapshots of the four source Sheets (+ this store) into archive/ on
    # the same private Shared Drive. The store itself is REPLACED every week, so
    # without this there is no copy of anything: a sheet deleted on Tuesday is
    # gone from the dashboard by the next Monday with no way back. Every source
    # is owned by a different person outside this repo - see archive.py.
    #
    # Deliberately AFTER the upload and never fatal: the dashboard refreshing is
    # what people actually depend on, and a Drive hiccup here must not cost them
    # that. Failures are reported, not raised.
    print("[8b] Archiving the source sheets …", flush=True)
    try:
        archive.run(svc, cfg, cfg["store_folder_id"],
                    when=datetime.now(IST).date(), store_bytes=store_bytes)
    except Exception as e:
        print(f"   WARNING: archive step failed entirely ({e})")

    print(f"Done — dashboard data refreshed "
          f"({datetime.now(IST).strftime('%d %b %Y, %H:%M IST')}).")


if __name__ == "__main__":
    main()

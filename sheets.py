"""
sheets.py — source adapters for the dashboard.

Two sources, both returning the same shape `data.build()` expects:
  • Google Sheets (live, default) via gspread with FORMATTED_VALUE, so COUNTIF
    formula cells come back as the computed "Present"/"Absent" string.
  • A local .xlsx (dev fallback) read with openpyxl.

Reuses the SAME service account + secrets as the rest of the app
(`[gcp_service_account]`, `[drive].roster_id`, `[drive].l2_id`).

Public API:
  load(source="sheets"|"xlsx", xlsx_path=None) -> (sheets, l2_lookup)
     sheets    = {tab_name: [[cell, ...], ...]}
     l2_lookup = {(batch_label, mm_dd): topic, mm_dd: topic}
"""
from __future__ import annotations
import re

from attendance_core import _mmdd

_SCOPES = ["https://www.googleapis.com/auth/spreadsheets.readonly",
           "https://www.googleapis.com/auth/drive.readonly"]


# ── L2 helpers ────────────────────────────────────────────────────────────────
def _l2_batch_labels(name: str) -> set:
    """Extract roster-style batch labels ('B17') from a messy L2 'Batch Name'
    like 'B19 Alpha Beta', 'B 22 IC', 'B 15-21 IC', 'B Jan2025(15-18)'."""
    s = str(name or "")
    nums: set[int] = set()
    for m in re.finditer(r"b\s*0*(\d{1,3})", s, re.I):       # B17, B 22, b7
        nums.add(int(m.group(1)))
    for m in re.finditer(r"(\d{1,3})\s*[-–]\s*(\d{1,3})", s):  # ranges 15-21, 15-18
        a, b = int(m.group(1)), int(m.group(2))
        if a <= b and b - a < 30:
            nums.update(range(a, b + 1))
    return {f"B{n}" for n in nums}


def _hdr_idx(header, *needles):
    """First header column whose lowercased text contains any needle."""
    low = [str(c or "").strip().lower() for c in header]
    for i, h in enumerate(low):
        if any(n in h for n in needles):
            return i
    return None


def build_l2_lookup(l2_tabs: dict[str, list[list]]) -> dict:
    """Build {(batch_label, mm_dd): topic, mm_dd: topic} from the L2 monthly tabs.

    Columns are located by header name *per tab* (their positions drift). A
    (batch, date) key is added whenever a batch is extractable; a date-only key
    is also kept as a shared-schedule fallback (first non-blank topic for a date).
    """
    lookup: dict = {}
    for rows in l2_tabs.values():
        if not rows:
            continue
        header = rows[0]
        di = _hdr_idx(header, "date")
        ti = _hdr_idx(header, "topic")
        bi = _hdr_idx(header, "batch name", "batch")
        if di is None or ti is None:
            continue
        for r in rows[1:]:
            topic = str(r[ti]).strip() if ti < len(r) and r[ti] is not None else ""
            mm = _mmdd(r[di]) if di < len(r) else None
            if not topic or not mm:
                continue
            lookup.setdefault(mm, topic)                     # shared fallback
            if bi is not None and bi < len(r):
                for lbl in _l2_batch_labels(r[bi]):
                    lookup[(lbl, mm)] = topic                 # batch-specific wins
    return lookup


# ── Google Sheets source (live) ───────────────────────────────────────────────
def _client():
    import streamlit as st
    import gspread
    from google.oauth2.service_account import Credentials
    info = dict(st.secrets["gcp_service_account"])
    creds = Credentials.from_service_account_info(info, scopes=_SCOPES)
    return gspread.authorize(creds)


def _spreadsheet_tabs(gc, spreadsheet_id):
    """{worksheet_title: get_all_values()} — FORMATTED_VALUE resolves formulas."""
    sh = gc.open_by_key(spreadsheet_id)
    return {ws.title: ws.get_all_values() for ws in sh.worksheets()}


def load_sheets():
    """(roster_tabs, l2_lookup) read live from Google Sheets."""
    import streamlit as st
    drive = st.secrets["drive"]
    gc = _client()
    roster = _spreadsheet_tabs(gc, drive["roster_id"])
    l2_lookup = {}
    if drive.get("l2_id"):
        l2_lookup = build_l2_lookup(_spreadsheet_tabs(gc, drive["l2_id"]))
    return roster, l2_lookup


# ── xlsx source (local dev fallback) ──────────────────────────────────────────
def _xlsx_tabs(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    out = {ws.title: [list(r) for r in ws.iter_rows(values_only=True)] for ws in wb.worksheets}
    wb.close()
    return out


def load_xlsx(roster_path, l2_path=None):
    """(roster_tabs, l2_lookup) read from local .xlsx files (offline dev)."""
    roster = _xlsx_tabs(roster_path)
    l2_lookup = build_l2_lookup(_xlsx_tabs(l2_path)) if l2_path else {}
    return roster, l2_lookup


def load(source: str = "sheets", roster_path=None, l2_path=None):
    """Single entry point. source='sheets' (live) or 'xlsx' (local dev)."""
    if source == "xlsx":
        return load_xlsx(roster_path, l2_path)
    return load_sheets()

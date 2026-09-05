"""
AttendaSync core engine (v2 — multi-batch aware).

Marks Zoom webinar attendance into a Master Batch Roster from a bulk attendee ZIP,
using the L2 weekly schedule (Webinar ID -> batch/date/topic).

Key behaviour:
  * A session shared by several batches (e.g. "AI CAP B2 + B3 + B4", "AI CAP B1 - AI CAP B22")
    is marked in EVERY covered batch that has a roster sheet — nothing is skipped.
  * If a webinar ID isn't in L2, the batch is recovered from the attendee's folder name.
  * If several attendee files land on the same batch+date, their attendee lists are merged.

Matching rule (default 'exact', per the auditing guide):
  Registered mail == Email (lower-cased, whitespace-stripped);
  Registered Number == Phone (digits only, trailing .0 removed, compared on the
  last 10 digits); blanks never match.
Optional 'inclusive' mode also uses the WhatsApp / broadcast columns.
"""
import re, csv, io, zipfile, datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ---------------- helpers ----------------
def _digits(s): return re.sub(r'\D', '', s or '')
def _wid(v):    return _digits(str(v)) if v is not None else ''

def _mmdd(v):
    if v is None: return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return f'{v.month:02d}_{v.day:02d}'
    s = str(v)
    m = re.search(r'(20\d\d)[_\-/](\d{1,2})[_\-/](\d{1,2})', s)
    if m: return f'{int(m.group(2)):02d}_{int(m.group(3)):02d}'
    months = dict(jan=1, feb=2, mar=3, apr=4, may=5, jun=6, jul=7,
                  aug=8, sep=9, oct=10, nov=11, dec=12)
    up = s.upper()
    for nm, mn in months.items():
        mt = re.search(r'([\dOo]{1,2})\s*(?:ST|ND|RD|TH)?\s*' + nm.upper(), up)
        if mt:
            return f'{mn:02d}_{int(mt.group(1).replace("O","0").replace("o","0")):02d}'
    return None

# A session column header is the date, plus the POD when the session was for
# one domain: "2026_08_23 | Techies". From B35 a single date carries up to
# eleven different sessions, so the date alone no longer identifies a column -
# keying on it merged all eleven attendee sets into one, which is how B35's
# 23 Aug came to report 1,193 present against a 916-member POD.
POD_SEP = " | "


def _col_pod(header):
    """Session column header -> its POD name, or '' for a whole-batch session."""
    s = str(header or "")
    return s.split(POD_SEP, 1)[1].strip() if POD_SEP in s else ""


def _col_header(ymd, pod):
    return f"{ymd}{POD_SEP}{pod}" if pod else ymd


def _parse_filename(fname):
    m = re.match(r'attendee_(\d+)_((20\d\d)_(\d{2})_(\d{2}))', fname)
    return (m.group(1), m.group(2)) if m else None

def _track(seg):
    s = seg.lower()
    # BSIAI first: its folders and L2 rows say 'BSIAI B1', 'BSI AI B3' or a
    # plain 'BSI B1'. Without this they fall through to CAP and collide head-on
    # with AI CAP B1/B2 - same key, different programme, silently wrong marks.
    if re.search('(?<![a-z])bsi', s):  return 'BSIAI'
    if 'ecap' in s or 'e-cap' in s: return 'ECAP'
    if re.search(r'\bus\b', s):      return 'US'
    if 'lcp' in s:                   return 'LCP'
    return 'CAP'

def extract_batches(text):
    """A batch-name string -> set of (track, number) keys. Handles &, + / commas, and - ranges."""
    keys = set()
    for seg in re.split(r'&|\band\b', text):
        tr = _track(seg)
        # 'CAP\s*B?' also catches the run-together form the POD rows use
        # ('AICAPB35, B36-Techies'): there is no word boundary before B35, so
        # the B pattern alone silently credits the session to B36 only.
        nums = sorted(set(int(x) for x in re.findall(r'\bB(\d{1,2})\b', seg, re.I)) |
                      set(int(x) for x in re.findall(r'CAP\s*B?(\d{1,2})\b', seg, re.I)))
        if not nums:
            continue
        is_list  = ('+' in seg) or (',' in seg)
        is_range = (not is_list) and bool(re.search(r'[-–]', seg)) and len(nums) >= 2
        chosen = range(min(nums), max(nums) + 1) if is_range else nums
        for n in chosen:
            keys.add((tr, n))
    return keys

def _sheet_key(name):
    """Roster sheet name -> (track, number) or None."""
    if 'att' in name.lower(): return None
    m = re.search(r'B?(\d{1,2})\b', name.split('-')[0])
    if not m: return None
    return (_track(name), int(m.group(1)))

def _folder_batches(member_path):
    """Recover batch keys from an attendee file's folder when the webinar isn't in L2."""
    folder = member_path.rsplit('/', 1)[0] if '/' in member_path else member_path
    rest = re.sub(r'^\d{4}-\d{2}-\d{2}\s*-\s*', '', folder)
    parts, batch_parts = re.split(r'\s+-\s+', rest), []
    for p in parts:                       # keep leading segments that name batches; stop at topic
        # Same run-together form extract_batches handles ('AICAPB37- Techies'):
        # without it this gate rejects the segment, the folder resolves to NO
        # batch, and the session is never even downloaded - L2 never gets asked.
        if re.search(r'\bB\d{1,2}\b|CAP\s*B?\d{1,2}\b', p, re.I): batch_parts.append(p)
        else: break
    return extract_batches(' - '.join(batch_parts)) if batch_parts else set()

# ---------------- L2 schedule ----------------
def parse_l2(l2_bytes, with_labels=False, with_mentors=False):
    """webinar_id -> (frozenset[(track,num)], topic). Includes combined/range sessions.

    With `with_labels`, also returns {webinar_id: raw 'Batch Name' cell} so callers
    can show a session the way L2 names it ('AI CAP B35 - Techies', 'AI CAP B8 + B22')
    rather than a bare parsed number. With `with_mentors`, also returns
    {webinar_id: 'Mentor' cell} - who actually taught it. Both extra maps are
    first-wins per webinar id, exactly like `out`, so they never disagree.

    The flags are separate rather than one wider return because five callers
    already unpack the two-tuple; adding a third element unconditionally would
    break every one of them."""
    wb = load_workbook(io.BytesIO(l2_bytes), data_only=True)
    out, labels, mentors = {}, {}, {}
    for ws in wb.worksheets:
        bcol = tcol = wcol = mcol = hrow = None
        for r in range(1, min(ws.max_row or 1, 6) + 1):
            for c in range(1, (ws.max_column or 1) + 1):
                v = str(ws.cell(r, c).value or '').strip().lower()
                if v == 'batch name':  bcol, hrow = c, r
                elif v == 'topic name': tcol = c
                elif v == 'webinar id': wcol = c
                # EXACT match only: the same header row also carries "Mentor's
                # email" and "In-house/Freelancer mentor", and a substring test
                # would bind the column to whichever came first.
                elif v == 'mentor': mcol = c
            if bcol and wcol: break
        if not (bcol and wcol and hrow): continue
        for r in range(hrow + 1, (ws.max_row or hrow) + 1):
            wid = _wid(ws.cell(r, wcol).value)
            if not wid: continue
            raw = str(ws.cell(r, bcol).value or '').strip()
            keys = extract_batches(raw)
            if not keys: continue
            topic = str(ws.cell(r, tcol).value or '').strip() if tcol else ''
            if wid not in out:
                out[wid] = (frozenset(keys), topic)
                labels[wid] = raw
                # Older monthly tabs have no Mentor column at all; absent simply
                # means "not recorded", which the UI renders as blank.
                if mcol:
                    mentors[wid] = str(ws.cell(r, mcol).value or '').strip()
    if with_labels and with_mentors: return out, labels, mentors
    if with_mentors: return out, mentors
    return (out, labels) if with_labels else out

# ---------------- Zoom attendee report ----------------
# Zoom exports two different shapes for the same webinar. Only one is usable.
#
#   Attendee Report  — has an `Attended` / `First Name` header row, and named
#                      sections ("Attendee Details", "Other Attended"). Carries
#                      Phone, which is ~9% of our roster matches.
#   flat list        — `Name (original name), Email, Join time, Leave time,
#                      Duration (minutes), Guest, ...`. No Attended column, no
#                      Phone. Zoom emits this for webinars run WITHOUT
#                      registration, so two sessions on the same day can differ.
#
# The flat shape used to parse to zero attendees WITHOUT raising, in both
# `parse_attendees` (here) and `day1_analysis.read_attendees`. That marked a
# whole batch absent, published a real-looking 0%, and left the run green.
# `is_attendee_report` is the single definition of "usable shape" that all three
# callers now gate on. Only `bsiai.sessions_from_files` ever guarded this.
# The gate is the PARSE RESULT, not the shape. Shape detection alone is not
# enough: a tab-delimited or UTF-16 report still carries a readable "Attendee
# Details" line, so it looks like a report, yet `parse_attendees` (comma dialect)
# reads zero rows out of it. Keying on "did we actually get anybody" catches the
# flat list, the wrong delimiter, a renamed column and whatever Zoom does next.
# `is_attendee_report` then only chooses which CAUSE to name in the warning.
ZERO_ATTENDEE_TAG = 'parsed to ZERO attendees'
UNREADABLE_FORMAT_TAG = 'not a Zoom Attendee Report'

_SECTION_MARKERS = {'attendee details', 'other attended',
                    'host details', 'panelist details'}


def zero_attendee_reason(text) -> str:
    """Why a file yielded nobody, phrased for the person who has to fix it.

    Two very different failures land here: a wrong export shape is fixed at the
    extractor, a report whose rows will not parse is usually a delimiter or an
    encoding problem. Saying which saves the triage."""
    if not is_attendee_report(text):
        return (f'{UNREADABLE_FORMAT_TAG} - no Attended/First Name header and no '
                f'Attendee Details section. A flat participant list? Re-export it '
                f'as an Attendee Report.')
    return ('it looks like an Attendee Report, but no row yielded an Email or '
            'Phone - wrong delimiter (tab, not comma), a non-UTF-8 encoding, or '
            'renamed columns.')


def is_attendee_report(text) -> bool:
    """True when `text` is a real Zoom *Attendee Report* rather than a flat
    participant list.

    Accepts EITHER marker, because the two parsers key on different ones: the
    marker here wants the `Attended` + `First Name` header row, while
    `day1_analysis.read_attendees` walks the named sections. A genuine report
    has both; the flat list has neither.

    A report whose header exists but carries NO data rows is a genuinely empty
    session (nobody joined), not a format problem — this returns True for it, so
    a real zero is still allowed to be zero. Scans with early exit rather than
    materialising the file, since attendee CSVs run to ~100 MB in memory.
    """
    if not text:
        return False
    for r in csv.reader(io.StringIO(text)):
        if not r:
            continue
        c0 = r[0].strip()
        if c0 == 'Attended' and 'First Name' in r:
            return True
        if c0.lower() in _SECTION_MARKERS:
            return True
    return False


def parse_attendees(text):
    rows = list(csv.reader(io.StringIO(text)))
    hidx = next((i for i, r in enumerate(rows)
                 if r and r[0].strip() == 'Attended' and 'First Name' in r), None)
    emails, ph_full, ph_last10 = set(), set(), set()
    if hidx is None: return emails, ph_full, ph_last10
    h = rows[hidx]
    # EITHER key is enough. Matching is an OR (email hit or phone hit), so a
    # report carrying only one of the two columns still marks everyone whose
    # roster row agrees on that one. Requiring both used to throw away a whole
    # readable file: a registration form that never collected a phone number has
    # no Phone column, `h.index('Phone')` raised, and the emails that WERE there
    # went in the bin — the session then looked like a broken export and was
    # skipped. Only a report with neither column is genuinely unusable.
    ei = h.index('Email') if 'Email' in h else None
    pi = h.index('Phone') if 'Phone' in h else None
    if ei is None and pi is None: return emails, ph_full, ph_last10
    for r in rows[hidx + 1:]:
        if not r: continue
        if r[0].strip() == 'Attended': break
        e = re.sub(r'\s', '', (r[ei] or '')).lower() if ei is not None and len(r) > ei else ''
        p = (_digits(re.sub(r'\.0$', '', (r[pi] or '').strip()))
             if pi is not None and len(r) > pi else '')
        if e: emails.add(e)
        if p:
            ph_full.add(p)
            if len(p) >= 10: ph_last10.add(p[-10:])
    return emails, ph_full, ph_last10

# ---------------- roster helpers ----------------
def _header_row(ws):
    for r in range(1, 4):
        j = ' '.join(str(ws.cell(r, c).value or '') for c in range(1, 13)).lower()
        if 'registered number' in j or 'registered mail' in j: return r
    return 1

def _col(ws, hr, *needles):           # first column whose header contains all needles
    for c in range(1, (ws.max_column or 1) + 1):
        h = str(ws.cell(hr, c).value or '').strip().lower()
        if all(nd in h for nd in needles): return c
    return None

def _last_used(ws):
    last = 10
    for c in range(1, (ws.max_column or 1) + 1):
        for r in range(1, (ws.max_row or 1) + 1):
            if ws.cell(r, c).value is not None:
                last = c; break
    return last

def _cell_email(v):
    if v is None: return ''
    s = str(v); return re.sub(r'\s', '', s).lower() if '@' in s else ''
def _phone_hit(p, ph_full, ph_last10) -> bool:
    """Does roster number `p` appear in a report's phone sets?

    Compared on the LAST 10 DIGITS, not the whole string. The two sources use
    different conventions and always have: measured 2026-09-06, 99.4% of roster
    `Registered Number` values carry the 91 country code (12 digits) while 78.8%
    of Zoom `Phone` values do not (10 digits). A full-string comparison therefore
    tested "919886440098" against "9886440098" and missed, which made the phone
    half of "email OR phone" dead in production — it added 1 to 6 students per
    session where the last-10 rule adds 126 to 145 (+8.4% to +9.5% attendance).

    Safe for this data: an Indian mobile number IS its last 10 digits, so two
    different people cannot collide on them. The full-string check is kept first
    for numbers too short to have a last-10 form. `day1_analysis.norm_phone` has
    always compared this way; this is the marker catching up, not a new rule.
    """
    if not p:
        return False
    if p in ph_full:
        return True
    return len(p) >= 10 and p[-10:] in ph_last10


def _cell_phone(v):
    if v is None: return ''
    if isinstance(v, (int, float)):
        s = str(int(v)) if float(v).is_integer() else str(v)
    else:
        s = re.sub(r'\.0$', '', str(v).strip())
    return _digits(s)

# ---------------- main entry ----------------
def process(roster_bytes, l2_bytes, zip_bytes, mode='exact', values_only=False):
    """Mark attendance from an uploaded ZIP of Zoom attendee CSVs.

    Thin wrapper: unpacks the ZIP into (name, raw_bytes) pairs and hands them to
    process_files(). Returns (output_xlsx_bytes, report_rows, warnings).
    """
    z = zipfile.ZipFile(io.BytesIO(zip_bytes))
    files = [(info, z.read(info)) for info in z.namelist() if not info.endswith('/')]
    return process_files(roster_bytes, l2_bytes, files, mode=mode, values_only=values_only)


def process_files(roster_bytes, l2_bytes, attendee_files, mode='exact', values_only=False):
    """Same engine as process(), but attendee data arrives as a list of
    (name, raw_bytes) pairs instead of a ZIP.

    This lets the marker run off ANY source — an uploaded ZIP, or files pulled
    live from a Google Drive folder/zip — without changing the matching logic.
    `name` should keep its folder path (e.g. "2025-05-09 - AI CAP B17 - Topic/
    attendee_123_2025_05_09.csv") so the folder-name fallback still works.
    `l2_bytes` may be falsy; then sessions are mapped by folder name only.

    `values_only`: some rosters store attendance as FORMULAS. With the default
    (False) we keep formulas (right when the output will be re-opened/recomputed
    in Excel/Sheets). Set True to load cached values instead — formulas become
    static text, so the marked workbook still carries every Present/Absent when
    read straight back by a values-only reader (the dashboard). Without this,
    saving a formula-based roster drops those cached values.

    Returns (output_xlsx_bytes, report_rows, warnings).
    """
    import pods

    wb = load_workbook(io.BytesIO(roster_bytes), data_only=values_only)
    wid_map, l2_labels = (parse_l2(l2_bytes, with_labels=True) if l2_bytes
                          else ({}, {}))

    key_sheet = {}
    for name in wb.sheetnames:
        k = _sheet_key(name)
        if k and k not in key_sheet: key_sheet[k] = name

    # (wid,ymd) -> [(name, bytes), ...]. The bytes ride along instead of sitting
    # in a name-keyed dict: the two Shared Drives use the SAME folder and file
    # naming for a shared session, so keying on the path made the second copy
    # overwrite the first and the good copy could be gone before any fallback
    # below could reach it.
    cands: dict = {}
    for name, data in attendee_files:
        pf = _parse_filename(name.split('/')[-1])
        if not pf: continue
        cands.setdefault(pf, []).append((name, data))

    # The same webinar is often uploaded twice: the two Shared Drives overlap
    # (166 sessions as of Sep 2026), and a single drive can hold "(1)" copies.
    # Only ONE copy is marked, so which one wins decides the numbers.
    #
    # Order of preference, highest first:
    #   1. the LAST caller-supplied copy that is not a "(1)" duplicate. Callers
    #      list drives in `attendee_folder_id` order, so the drive named last —
    #      "Zoom extracts", the authoritative one per the owner, 2026-09-06 —
    #      wins. This used to fall out of a last-one-wins dict write; it is
    #      explicit now, because swapping the ids in the secret silently
    #      reversed it.
    #   2. "(1)" copies, only if nothing else is left.
    #
    # Preference alone is not enough: the preferred copy can be the unreadable
    # export shape. So callers below try candidates in this order and take the
    # first that yields anybody, rather than skipping a session outright when a
    # perfectly good copy exists on the other drive.
    chosen = {}
    for pf, items in cands.items():
        chosen[pf] = ([c for c in reversed(items) if '(1)' not in c[0]]
                      + [c for c in reversed(items) if '(1)' in c[0]])

    # accumulate attendee sets per (sheet_name, mmdd, POD), merging multiple files
    acc, warnings = {}, []
    unknown_pods = set()
    for (wid, ymd), ranked in chosen.items():
        info = ranked[0][0]
        keys, topic = wid_map.get(wid, (None, ''))
        if keys is None:                                 # not in L2 -> recover from folder
            for cand, _data in ranked:                   # drives may name folders differently
                keys = _folder_batches(cand)
                if keys:
                    warnings.append(f'Webinar {wid}: not in L2 — batch taken from folder name')
                    break
        if not keys:
            warnings.append(f'Webinar {wid}: no batch info in L2 or folder — skipped ({info.split("/")[-1]})')
            continue

        # Take the first copy that yields anybody. A session is only skipped when
        # EVERY copy is unreadable, so an unusable export on the preferred drive
        # can no longer cost a session that the other drive holds intact.
        em = pf = p10 = None
        tried = []
        for cand, data in ranked:
            try:
                text = data.decode('utf-8-sig', errors='replace')
            except Exception as e:
                tried.append(f'{cand.split("/")[-1]}: {e}')
                continue
            c_em, c_pf, c_p10 = parse_attendees(text)
            if c_em or c_pf:
                info, em, pf, p10 = cand, c_em, c_pf, c_p10
                break
            tried.append(f'{cand.split("/")[-1]} - {zero_attendee_reason(text)}')

        # SKIP rather than mark. A file we cannot read yields nobody, and marking
        # that session would write Absent against every student in the batch —
        # indistinguishable, on the dashboard, from a session nobody attended.
        # Skipping means the column is never created, so the session is visibly
        # MISSING instead of silently wrong. pipeline.py turns this warning into a
        # refusal to publish. A webinar that genuinely had no attendees is caught
        # by the same net; that is deliberate, since it is far rarer than a broken
        # export and publishing a whole batch as absent is the worse error.
        if em is None and pf is None:
            warnings.append(
                f'Webinar {wid}: {ZERO_ATTENDEE_TAG} in all {len(ranked)} copy(ies) '
                f'- {"; ".join(tried)}. Session SKIPPED rather than marking the '
                f'whole batch absent.')
            continue
        mm = f'{int(ymd[5:7]):02d}_{int(ymd[8:10]):02d}'

        # Which POD is this session for? L2's own label is authoritative; the
        # folder name is the fallback, exactly as it is for the batch itself.
        # An unrecognised domain becomes a whole-batch session AND warns - it
        # must never be silently folded into another POD's denominator.
        pod = pods.from_l2_label(l2_labels.get(wid)) if l2_labels.get(wid) else None
        if pod is None:
            pod = pods.from_folder(info)
        if pod is None and l2_labels.get(wid):
            tail = str(l2_labels.get(wid)).split('-')[-1].strip()
            if tail and tail not in unknown_pods:
                unknown_pods.add(tail)
                warnings.append(f'Webinar {wid}: POD "{tail}" not recognised - '
                                f'counted against the whole batch. Add it to pods._ALIASES.')
        pod = '' if pod in (None, pods.WHOLE_BATCH) else pod
        targets = [key_sheet[k] for k in keys if k in key_sheet]
        if not targets:
            blist = ', '.join(f'B{n}' if t == 'CAP' else f'{t} B{n}' for t, n in sorted(keys))
            warnings.append(f'Webinar {wid}: batch(es) {blist} not in this roster — left empty (sheets are never created)')
            continue
        for sheet in targets:
            a = acc.setdefault((sheet, mm, pod), dict(ymd=ymd, topic=topic, pod=pod,
                                                      em=set(), pf=set(), p10=set()))
            a['em'] |= em; a['pf'] |= pf; a['p10'] |= p10
            if not a['topic']: a['topic'] = topic

    # write per sheet
    report = []
    sheets = sorted({k[0] for k in acc}, key=lambda n: _sheet_key(n) or ('', 0))
    for sheet in sheets:
        ws = wb[sheet]; hr = _header_row(ws)
        rm = _col(ws, hr, 'registered', 'mail')
        rn = _col(ws, hr, 'registered', 'number')
        wa = _col(ws, hr, 'whatsa');  bc = _col(ws, hr, 'broadcast')
        pc = _col(ws, hr, 'pod')      # 'POD pref' / 'POD Pref' / 'POD Prefrence'
        if rm is None and rn is None:
            warnings.append(f'Sheet "{sheet}": no Registered mail/number columns — skipped')
            continue
        dmap = {}
        for c in range(11, (ws.max_column or 1) + 1):
            h1 = ws.cell(1, c).value
            d = _mmdd(h1) or (_mmdd(ws.cell(2, c).value) if hr == 2 else None)
            if d: dmap.setdefault((d, _col_pod(h1)), c)
        nxt = _last_used(ws) + 1
        # Real rosters sometimes have merged cells in the body; writing a mark into
        # a merged (non-anchor) cell raises in openpyxl. Unmerge any range that
        # touches a data row (header merges in rows 1–hr are left intact, and dmap
        # above was already built from the headers, so this is safe).
        for rng in list(ws.merged_cells.ranges):
            if rng.max_row > hr:
                ws.unmerge_cells(str(rng))
        for (s, mm, pod) in sorted((k for k in acc if k[0] == sheet),
                                   key=lambda k: (k[1], k[2])):
            a = acc[(s, mm, pod)]
            if (mm, pod) in dmap:
                ci, kind = dmap[(mm, pod)], 're-mark'
            else:
                ci, kind = nxt, 'NEW'; nxt += 1
                ws.cell(1, ci).value = _col_header(a['ymd'], pod)
                if hr == 2 and a['topic']: ws.cell(2, ci).value = a['topic']
                dmap[(mm, pod)] = ci
            em, pf, p10 = a['em'], a['pf'], a['p10']
            pres = tot = outside = 0
            for r in range(hr + 1, (ws.max_row or hr) + 1):
                # A POD session is only for that POD. Marking everyone else
                # 'Absent' is simply false - they were never invited - and it put
                # ten wrong Absents against every student in the downloadable
                # roster the moment a batch ran eleven PODs in a day.
                if pod and pc:
                    rp, _multi = pods.from_roster_cell(ws.cell(r, pc).value)
                    if (rp or pods.UNKNOWN) != pod:
                        # still worth knowing if they turned up
                        e0 = _cell_email(ws.cell(r, rm).value) if rm else ''
                        p0 = _cell_phone(ws.cell(r, rn).value) if rn else ''
                        if (e0 and e0 in em) or _phone_hit(p0, pf, p10):
                            outside += 1
                        continue
                if mode == 'exact':
                    e = _cell_email(ws.cell(r, rm).value) if rm else ''
                    p = _cell_phone(ws.cell(r, rn).value) if rn else ''
                    if not e and not p: continue
                    hit = (e and e in em) or _phone_hit(p, pf, p10)
                else:
                    es = [x for x in (_cell_email(ws.cell(r, c).value) for c in (rm, bc) if c) if x]
                    ps = [x for x in (_cell_phone(ws.cell(r, c).value) for c in (rn, wa) if c) if x]
                    if not es and not ps: continue
                    hit = any(x in em for x in es) or any(len(x) >= 10 and x[-10:] in p10 for x in ps)
                tot += 1
                ws.cell(r, ci).value = 'Present' if hit else 'Absent'
                if hit: pres += 1
            # A "POD" session most of the batch attended is not a POD session -
            # either L2 mislabels a whole-batch session, or the domain is wrong.
            # Loudly, because the numbers look plausible either way.
            if pod and outside > max(20, pres):
                warnings.append(
                    f'{sheet} {a["ymd"]} "{pod}": {outside} attendee(s) matched '
                    f'students OUTSIDE that POD (vs {pres} inside) - is this really '
                    f'a POD session, or does L2 label a whole-batch session with a '
                    f'domain?')
            report.append(dict(batch=sheet, date=a['ymd'], col=get_column_letter(ci),
                               kind=kind, topic=a['topic'], pod=pod,
                               present=pres, total=tot, outside=outside))
    out = io.BytesIO(); wb.save(out)
    return out.getvalue(), report, warnings
